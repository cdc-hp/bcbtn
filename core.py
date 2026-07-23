from __future__ import annotations

import base64
import csv
from contextlib import contextmanager
from difflib import SequenceMatcher
import hashlib
import hmac
import json
import os
import re
import shutil
import sqlite3
import sys
import threading
import time
import unicodedata
import uuid
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable, Sequence

from openpyxl import Workbook, load_workbook

import backup_manager
from duplicate_config import CASE_CRITERIA_LABELS, CaseDuplicateCriteria, load_case_criteria, load_rules

APP_NAME = "Giám sát dịch bệnh"
# Không tự đồng bộ với VERSION.txt (dùng cho tên gói cài) — nhớ cập nhật cả 2 mỗi lần bump
# phiên bản, nếu không update_manager.is_newer_version() so sánh sai (từng lệch: VERSION.txt
# đã ở 0.6.0 nhưng hằng số này còn 0.5.0).
VERSION = "0.7.0"


def _base_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


BASE_DIR = _base_dir()


def _user_data_root() -> Path:
    """Thư mục dữ liệu tách khỏi bộ cài để cài mới/cập nhật không đụng CSDL."""
    override = os.environ.get("GIAM_SAT_DICH_BENH_DATA_DIR", "").strip()
    if override:
        return Path(override).expanduser().resolve()
    if os.name == "nt":
        local_app_data = os.environ.get("LOCALAPPDATA")
        if local_app_data:
            return Path(local_app_data) / "CDC_HaiPhong" / "GiamSatDichBenh"
    return Path.home() / ".giam_sat_dich_benh"


USER_DATA_DIR = _user_data_root()
DATA_DIR = USER_DATA_DIR / "data"
BACKUP_DIR = USER_DATA_DIR / "backups"
UPDATE_CACHE_DIR = USER_DATA_DIR / "update_cache"
QUEUE_DIR = USER_DATA_DIR / "queue_uploads"
DB_PATH = DATA_DIR / "giam_sat_dich_benh.db"
_DB_INIT_LOCK = threading.RLock()

for directory in (DATA_DIR, BACKUP_DIR, UPDATE_CACHE_DIR, QUEUE_DIR):
    directory.mkdir(parents=True, exist_ok=True)

QUEUE_STATUSES = {"cho_nhap", "dang_nhap", "da_nhap", "loi"}
QUEUE_SOURCES = {"server_chinh", "server_phu"}


CASE_FIELDS: list[tuple[str, str]] = [
    ("STT", "source_stt"),
    ("Mã số", "case_code"),
    ("Họ tên", "full_name"),
    ("Ngày sinh", "birth_date_raw"),
    ("Nghề nghiệp", "occupation"),
    ("Nơi làm việc/ Học tập", "workplace"),
    ("Tỉnh (Nơi làm việc)", "workplace_province"),
    ("Xã (Nơi làm việc)", "workplace_commune"),
    ("Thôn (Nơi làm việc)", "workplace_village"),
    ("Dân tộc", "ethnicity"),
    ("Giới tính", "gender"),
    ("CMND", "national_id"),
    ("Điện thoại", "phone"),
    ("Nơi ở hiện nay", "current_address"),
    ("Tỉnh", "province"),
    ("Xã", "commune"),
    ("Thôn", "village"),
    ("Kinh độ", "longitude"),
    ("Vĩ độ", "latitude"),
    ("Chẩn đoán chính", "main_diagnosis"),
    ("Phân độ bệnh", "severity"),
    ("Chẩn đoán bệnh kèm theo", "comorbidity_diagnosis"),
    ("Tình trạng tiêm chủng", "vaccination_status"),
    ("Số lần tiêm, uống", "vaccination_doses"),
    ("Phân loại chẩn đoán", "diagnosis_classification"),
    ("Lấy mẫu xét nghiệm", "sample_collected"),
    ("Ngày lấy mẫu", "sample_date"),
    ("Đơn vị xét nghiệm", "lab_unit"),
    ("Loại xét nghiệm", "test_type"),
    ("Kết quả", "test_result"),
    ("Ghi chú", "notes"),
    ("Tình trạng hiện nay", "current_status"),
    ("Tiền sử dịch tễ", "epidemiological_history"),
    ("Ngày khởi phát", "onset_date"),
    ("Ngày nhập viện", "admission_date"),
    ("Ngày ra viện/tử vong", "discharge_or_death_date"),
    ("Biến chứng", "complications"),
    ("Tên người báo cáo", "reporter_name"),
    ("SĐT người báo cáo", "reporter_phone"),
    ("Email người báo cáo", "reporter_email"),
    ("Đơn vị báo cáo", "reporting_unit"),
    ("Tỉnh đơn vị báo cáo", "reporting_province"),
    ("Cơ sở điều trị", "treatment_facility"),
    ("Trạng thái", "record_status"),
    ("Thời gian báo cáo", "report_datetime"),
    ("Chẩn đoán thay đổi gần nhất", "latest_diagnosis_change"),
    ("Tình trạng thay đổi gần nhất", "latest_status_change"),
    ("Thời gian sửa", "modified_datetime"),
]

OUTBREAK_FIELDS: list[tuple[str, str]] = [
    ("STT", "source_stt"),
    ("Tên bệnh", "disease"),
    ("Địa điểm xảy ra ổ dịch", "location"),
    ("Ngày khởi phát trường hợp bệnh đầu tiên", "first_onset_date"),
    ("Ngày ổ dịch kết thúc hoạt động", "end_date"),
    ("Trạng thái", "status"),
    ("Số ca mắc", "case_count"),
    ("Số ca tử vong", "death_count"),
    ("Số mẫu XN", "sample_count"),
    ("Số mẫu (+)", "positive_count"),
    ("Ngày báo cáo", "report_datetime"),
    ("Đơn vị báo cáo", "reporting_unit"),
    ("Tỉnh báo cáo", "reporting_province"),
    ("Ngày nhận báo cáo ổ dịch bệnh đầu tiên", "first_report_received_date"),
    ("Ngày khởi phát trường hợp bệnh cuối cùng", "last_onset_date"),
]

CASE_LABELS = {db: label for label, db in CASE_FIELDS}
OUTBREAK_LABELS = {db: label for label, db in OUTBREAK_FIELDS}

DATE_FIELDS = {
    "sample_date",
    "onset_date",
    "admission_date",
    "discharge_or_death_date",
    "first_onset_date",
    "end_date",
    "first_report_received_date",
    "last_onset_date",
}
DATETIME_FIELDS = {"report_datetime", "modified_datetime"}
INTEGER_FIELDS = {"source_stt", "case_count", "death_count", "sample_count", "positive_count"}
FLOAT_FIELDS = {"longitude", "latitude"}


@dataclass
class ImportSummary:
    file_name: str
    entity_type: str
    detected_sheet: str
    rows_read: int = 0
    inserted: int = 0
    duplicates: int = 0
    skipped: int = 0
    issues: int = 0

    def as_text(self) -> str:
        kind = "ca bệnh" if self.entity_type == "case" else "ổ dịch"
        return (
            f"{self.file_name} — {kind}: đọc {self.rows_read}, thêm {self.inserted}, "
            f"trùng {self.duplicates}, bỏ qua {self.skipped}, cảnh báo {self.issues}."
        )


@contextmanager
def _connect(db_path: Path | str = DB_PATH):
    """Mở kết nối SQLite theo phạm vi và luôn đóng file khi rời khối with.

    sqlite3.Connection.__exit__ chỉ commit/rollback chứ không đóng kết nối;
    trên Windows điều đó giữ khóa file .db và làm sao lưu/xóa thất bại.
    """
    path = Path(db_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(path, timeout=15)
    conn.row_factory = sqlite3.Row
    try:
        conn.execute("PRAGMA foreign_keys = ON")
        conn.execute("PRAGMA busy_timeout = 10000")
        conn.execute("PRAGMA synchronous = NORMAL")
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def _ensure_column(conn: sqlite3.Connection, table: str, column: str, definition: str) -> None:
    existing = {str(row[1]) for row in conn.execute(f"PRAGMA table_info({table})").fetchall()}
    if column not in existing:
        conn.execute(f"ALTER TABLE {table} ADD COLUMN {column} {definition}")


def init_db(db_path: Path | str = DB_PATH) -> None:
    with _DB_INIT_LOCK:
        with _connect(db_path) as conn:
            conn.execute("PRAGMA journal_mode = WAL")
            conn.executescript(
                """
                CREATE TABLE IF NOT EXISTS cases (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    source_stt INTEGER,
                    case_code TEXT,
                    full_name TEXT,
                    birth_date_raw TEXT,
                    birth_year INTEGER,
                    occupation TEXT,
                    workplace TEXT,
                    workplace_province TEXT,
                    workplace_commune TEXT,
                    workplace_village TEXT,
                    ethnicity TEXT,
                    gender TEXT,
                    national_id TEXT,
                    phone TEXT,
                    current_address TEXT,
                    province TEXT,
                    commune TEXT,
                    village TEXT,
                    longitude REAL,
                    latitude REAL,
                    main_diagnosis TEXT,
                    severity TEXT,
                    comorbidity_diagnosis TEXT,
                    vaccination_status TEXT,
                    vaccination_doses TEXT,
                    diagnosis_classification TEXT,
                    sample_collected TEXT,
                    sample_date TEXT,
                    lab_unit TEXT,
                    test_type TEXT,
                    test_result TEXT,
                    notes TEXT,
                    current_status TEXT,
                    epidemiological_history TEXT,
                    onset_date TEXT,
                    admission_date TEXT,
                    discharge_or_death_date TEXT,
                    complications TEXT,
                    reporter_name TEXT,
                    reporter_phone TEXT,
                    reporter_email TEXT,
                    reporting_unit TEXT,
                    reporting_province TEXT,
                    treatment_facility TEXT,
                    record_status TEXT,
                    report_datetime TEXT,
                    latest_diagnosis_change TEXT,
                    latest_status_change TEXT,
                    modified_datetime TEXT,
                    source_file TEXT NOT NULL,
                    source_sheet TEXT,
                    source_row INTEGER,
                    row_hash TEXT NOT NULL UNIQUE,
                    imported_at TEXT NOT NULL,
                    raw_json TEXT
                );
    
                CREATE TABLE IF NOT EXISTS outbreaks (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    source_stt INTEGER,
                    disease TEXT,
                    location TEXT,
                    admin_area TEXT,
                    first_onset_date TEXT,
                    end_date TEXT,
                    status TEXT,
                    case_count INTEGER DEFAULT 0,
                    death_count INTEGER DEFAULT 0,
                    sample_count INTEGER DEFAULT 0,
                    positive_count INTEGER DEFAULT 0,
                    report_datetime TEXT,
                    reporting_unit TEXT,
                    reporting_province TEXT,
                    first_report_received_date TEXT,
                    last_onset_date TEXT,
                    source_file TEXT NOT NULL,
                    source_sheet TEXT,
                    source_row INTEGER,
                    row_hash TEXT NOT NULL UNIQUE,
                    imported_at TEXT NOT NULL,
                    raw_json TEXT
                );
    
                CREATE TABLE IF NOT EXISTS import_batches (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    file_name TEXT NOT NULL,
                    file_path TEXT,
                    entity_type TEXT NOT NULL,
                    sheet_name TEXT,
                    rows_read INTEGER DEFAULT 0,
                    inserted INTEGER DEFAULT 0,
                    duplicates INTEGER DEFAULT 0,
                    skipped INTEGER DEFAULT 0,
                    issue_count INTEGER DEFAULT 0,
                    imported_at TEXT NOT NULL
                );
    
                CREATE TABLE IF NOT EXISTS data_quality_issues (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    entity_type TEXT NOT NULL,
                    entity_id INTEGER,
                    source_file TEXT,
                    source_row INTEGER,
                    severity TEXT NOT NULL,
                    issue_type TEXT NOT NULL,
                    description TEXT,
                    created_at TEXT NOT NULL
                );
    
                CREATE TABLE IF NOT EXISTS duplicate_actions (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    entity_type TEXT NOT NULL,
                    keep_id INTEGER NOT NULL,
                    removed_ids_json TEXT NOT NULL,
                    backup_file TEXT,
                    created_at TEXT NOT NULL,
                    action_type TEXT DEFAULT 'merge',
                    merged_values_json TEXT,
                    restored_at TEXT
                );
    
                CREATE TABLE IF NOT EXISTS duplicate_trash (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    action_id INTEGER NOT NULL,
                    entity_type TEXT NOT NULL,
                    original_id INTEGER NOT NULL,
                    record_json TEXT NOT NULL,
                    deleted_at TEXT NOT NULL,
                    restored_at TEXT,
                    restored_id INTEGER
                );
    
                CREATE TABLE IF NOT EXISTS import_queue (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    commune TEXT NOT NULL,
                    week TEXT NOT NULL,
                    file_name TEXT NOT NULL,
                    file_path TEXT NOT NULL,
                    source TEXT NOT NULL DEFAULT 'server_chinh',
                    status TEXT NOT NULL DEFAULT 'cho_nhap',
                    submitted_by TEXT,
                    received_at TEXT NOT NULL,
                    imported_at TEXT,
                    import_batch_id INTEGER,
                    error_message TEXT,
                    archived_at TEXT
                );

                CREATE INDEX IF NOT EXISTS idx_import_queue_status ON import_queue(status);
                CREATE INDEX IF NOT EXISTS idx_import_queue_commune ON import_queue(commune, week);

                CREATE TABLE IF NOT EXISTS commune_accounts (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    commune TEXT NOT NULL UNIQUE,
                    username TEXT NOT NULL UNIQUE,
                    password_hash TEXT NOT NULL,
                    display_name TEXT,
                    active INTEGER NOT NULL DEFAULT 1,
                    created_at TEXT NOT NULL,
                    last_login_at TEXT
                );

                CREATE TABLE IF NOT EXISTS cdc_accounts (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    username TEXT NOT NULL UNIQUE,
                    password_hash TEXT NOT NULL,
                    display_name TEXT,
                    active INTEGER NOT NULL DEFAULT 1,
                    created_at TEXT NOT NULL,
                    last_login_at TEXT
                );

                CREATE TABLE IF NOT EXISTS audit_log (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    created_at TEXT NOT NULL,
                    actor TEXT NOT NULL,
                    action TEXT NOT NULL,
                    commune TEXT,
                    detail TEXT
                );

                CREATE INDEX IF NOT EXISTS idx_audit_log_action ON audit_log(action, created_at);
                CREATE INDEX IF NOT EXISTS idx_audit_log_commune ON audit_log(commune, created_at);
                CREATE INDEX IF NOT EXISTS idx_duplicate_trash_action ON duplicate_trash(action_id, restored_at);
                CREATE INDEX IF NOT EXISTS idx_cases_name ON cases(full_name);
                CREATE INDEX IF NOT EXISTS idx_cases_code ON cases(case_code);
                CREATE INDEX IF NOT EXISTS idx_cases_diag ON cases(main_diagnosis);
                CREATE INDEX IF NOT EXISTS idx_cases_onset ON cases(onset_date);
                CREATE INDEX IF NOT EXISTS idx_cases_commune ON cases(commune);
                CREATE INDEX IF NOT EXISTS idx_outbreaks_disease ON outbreaks(disease);
                CREATE INDEX IF NOT EXISTS idx_outbreaks_status ON outbreaks(status);
                CREATE INDEX IF NOT EXISTS idx_outbreaks_onset ON outbreaks(first_onset_date);
                CREATE INDEX IF NOT EXISTS idx_outbreaks_admin ON outbreaks(admin_area);
                CREATE INDEX IF NOT EXISTS idx_quality_type ON data_quality_issues(entity_type, severity);
                """
            )
            _ensure_column(conn, "duplicate_actions", "action_type", "TEXT DEFAULT 'merge'")
            _ensure_column(conn, "duplicate_actions", "merged_values_json", "TEXT")
            _ensure_column(conn, "duplicate_actions", "restored_at", "TEXT")
            _ensure_column(conn, "import_queue", "archived_at", "TEXT")

def strip_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="minutes")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_key(value: Any) -> str:
    text = strip_text(value).lower()
    text = text.replace("đ", "d")
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def normalize_header(value: Any) -> str:
    """Chuẩn hóa tiêu đề Excel, chịu được xuống dòng, dấu *, /, ngoặc và đánh số cột."""
    text = normalize_key(value).replace("\n", " ").replace("\r", " ")
    text = re.sub(r"^\s*\d+[\.\)]\s*", "", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def parse_date_value(value: Any, with_time: bool = False) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="minutes") if with_time else value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = strip_text(value)
    fmts = (
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%H:%M %d/%m/%Y",
        "%d/%m/%Y",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
    )
    for fmt in fmts:
        try:
            dt = datetime.strptime(text, fmt)
            return dt.isoformat(sep=" ", timespec="minutes") if with_time else dt.date().isoformat()
        except ValueError:
            continue
    return text


def parse_int(value: Any) -> int:
    if value is None or value == "":
        return 0
    try:
        return int(float(str(value).replace(",", ".")))
    except (ValueError, TypeError):
        return 0


def parse_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(str(value).replace(",", "."))
    except (ValueError, TypeError):
        return None


def extract_birth_year(value: Any) -> int | None:
    text = strip_text(value)
    years = re.findall(r"(?:19|20)\d{2}", text)
    if years:
        return int(years[-1])
    return None


def extract_admin_area(location: str) -> str:
    if not location:
        return ""
    parts = [p.strip() for p in re.split(r"\s+-\s+", location) if p.strip()]
    if len(parts) >= 2:
        if "hải phòng" in normalize_key(parts[-1]) and len(parts) >= 2:
            return parts[-2]
        return parts[-1]
    patterns = re.findall(r"(?:Phường|Xã|Đặc khu|Thị trấn)\s+[^,;]+", location, flags=re.I)
    return patterns[-1].strip() if patterns else ""


def _row_hash(entity_type: str, payload: dict[str, Any]) -> str:
    canonical = {k: payload.get(k, "") for k in sorted(payload) if k not in {"source_file", "source_sheet", "source_row"}}
    raw = json.dumps(canonical, ensure_ascii=False, sort_keys=True, default=str).encode("utf-8")
    return hashlib.sha256(entity_type.encode("utf-8") + b"|" + raw).hexdigest()


CASE_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "source_stt": ("số thứ tự", "stt."),
    "case_code": ("mã ca bệnh", "mã bệnh nhân", "mã trường hợp", "mã số ca bệnh"),
    "full_name": ("họ và tên", "tên bệnh nhân", "họ tên bệnh nhân"),
    "birth_date_raw": ("năm sinh", "ngày tháng năm sinh"),
    "national_id": ("cccd", "cmnd cccd", "số cmnd", "số cccd", "cccd cmnd"),
    "phone": ("số điện thoại", "sđt", "điện thoại liên hệ"),
    "current_address": ("địa chỉ hiện nay", "nơi cư trú hiện nay", "địa chỉ nơi ở hiện nay"),
    "province": ("tỉnh thành phố", "tỉnh tp", "tỉnh nơi ở"),
    "commune": ("phường xã", "xã phường", "xã nơi ở"),
    "village": ("thôn tổ", "thôn xóm", "tổ dân phố"),
    "main_diagnosis": ("tên bệnh", "chẩn đoán", "bệnh chẩn đoán chính"),
    "onset_date": ("ngày phát bệnh", "ngày bắt đầu khởi phát"),
    "report_datetime": ("ngày báo cáo", "thời điểm báo cáo", "ngày giờ báo cáo"),
    "reporting_unit": ("cơ quan báo cáo", "đơn vị gửi báo cáo"),
    "record_status": ("trạng thái bản ghi", "tình trạng bản ghi"),
}

OUTBREAK_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "source_stt": ("số thứ tự", "stt."),
    "disease": ("bệnh", "tên dịch bệnh"),
    "location": ("địa điểm ổ dịch", "nơi xảy ra ổ dịch", "địa chỉ ổ dịch"),
    "first_onset_date": ("ngày khởi phát ca đầu tiên", "ngày khởi phát trường hợp đầu tiên"),
    "end_date": ("ngày kết thúc ổ dịch", "ngày ổ dịch kết thúc"),
    "case_count": ("số mắc", "tổng số ca mắc"),
    "death_count": ("số tử vong", "tổng số ca tử vong"),
    "sample_count": ("số mẫu xét nghiệm", "tổng số mẫu xn"),
    "positive_count": ("số mẫu dương tính", "số mẫu dương", "số mẫu xn dương tính"),
    "report_datetime": ("thời gian báo cáo", "ngày giờ báo cáo"),
    "reporting_province": ("tỉnh đơn vị báo cáo", "tỉnh thành báo cáo"),
    "first_report_received_date": ("ngày nhận báo cáo đầu tiên", "ngày nhận báo cáo ổ dịch đầu tiên"),
    "last_onset_date": ("ngày khởi phát ca cuối cùng", "ngày khởi phát trường hợp cuối cùng"),
}


def _make_header_map(fields: Sequence[tuple[str, str]], aliases: dict[str, tuple[str, ...]]) -> dict[str, str]:
    result: dict[str, str] = {}
    for label, db_field in fields:
        result[normalize_header(label)] = db_field
        for alias in aliases.get(db_field, ()):
            result[normalize_header(alias)] = db_field
    return result


def _mapping_for_row(row: Sequence[Any], header_map: dict[str, str]) -> dict[int, str]:
    mapping: dict[int, str] = {}
    used_fields: set[str] = set()
    for idx, value in enumerate(row):
        key = normalize_header(value)
        field = header_map.get(key)
        if field and field not in used_fields:
            mapping[idx] = field
            used_fields.add(field)
    return mapping


def _find_header_row(ws) -> tuple[str, int, dict[int, str]] | None:
    case_map = _make_header_map(CASE_FIELDS, CASE_HEADER_ALIASES)
    outbreak_map = _make_header_map(OUTBREAK_FIELDS, OUTBREAK_HEADER_ALIASES)
    best: tuple[int, str, int, dict[int, str]] | None = None

    # Một số hệ thống xuất XLSX khai báo dimension=A1 dù thực tế có nhiều cột/dòng.
    # reset_dimensions buộc openpyxl đọc theo dữ liệu XML thực tế.
    reset = getattr(ws, "reset_dimensions", None)
    if callable(reset):
        reset()

    for row_no, row in enumerate(ws.iter_rows(min_row=1, max_row=60, values_only=True), start=1):
        case_mapping = _mapping_for_row(row, case_map)
        outbreak_mapping = _mapping_for_row(row, outbreak_map)
        case_fields = set(case_mapping.values())
        outbreak_fields = set(outbreak_mapping.values())

        case_core = len(case_fields & {
            "case_code", "full_name", "birth_date_raw", "main_diagnosis",
            "onset_date", "report_datetime", "reporting_unit"
        })
        outbreak_core = len(outbreak_fields & {
            "disease", "location", "first_onset_date", "case_count",
            "report_datetime", "reporting_unit"
        })

        if "full_name" in case_fields and case_core >= 3 and len(case_fields) >= 6:
            candidate = (len(case_fields), "case", row_no, case_mapping)
            if best is None or candidate[0] > best[0]:
                best = candidate
        if {"disease", "location"}.issubset(outbreak_fields) and outbreak_core >= 4 and len(outbreak_fields) >= 5:
            candidate = (len(outbreak_fields), "outbreak", row_no, outbreak_mapping)
            if best is None or candidate[0] > best[0]:
                best = candidate

    if best:
        _, entity_type, row_no, mapping = best
        return entity_type, row_no, mapping
    return None


def detect_excel(path: Path | str) -> tuple[str, str, int, dict[int, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    diagnostics: list[str] = []
    try:
        for ws in wb.worksheets:
            found = _find_header_row(ws)
            if found:
                entity_type, header_row, mapping = found
                return entity_type, ws.title, header_row, mapping
            diagnostics.append(ws.title)
    finally:
        wb.close()
    sheets = ", ".join(diagnostics) if diagnostics else "không có trang tính"
    raise ValueError(
        "Không nhận diện được cấu trúc file ca bệnh hoặc ổ dịch. "
        f"Đã kiểm tra: {sheets}. Hãy bảo đảm file có hàng tiêu đề chứa các cột như "
        "Họ tên/Mã số/Ngày khởi phát hoặc Tên bệnh/Địa điểm xảy ra ổ dịch."
    )


def _normalize_payload(entity_type: str, payload: dict[str, Any]) -> dict[str, Any]:
    clean: dict[str, Any] = {}
    for key, value in payload.items():
        if key in DATE_FIELDS:
            clean[key] = parse_date_value(value, with_time=False)
        elif key in DATETIME_FIELDS:
            clean[key] = parse_date_value(value, with_time=True)
        elif key in INTEGER_FIELDS:
            clean[key] = parse_int(value)
        elif key in FLOAT_FIELDS:
            clean[key] = parse_float(value)
        else:
            clean[key] = strip_text(value)
    if entity_type == "case":
        clean["birth_year"] = extract_birth_year(clean.get("birth_date_raw"))
    else:
        clean["admin_area"] = extract_admin_area(clean.get("location", ""))
    return clean


def _is_empty_payload(entity_type: str, payload: dict[str, Any]) -> bool:
    if entity_type == "case":
        fields = ("case_code", "full_name", "main_diagnosis", "onset_date", "report_datetime")
    else:
        fields = ("disease", "location", "first_onset_date", "report_datetime")
    return not any(payload.get(k) not in (None, "", 0) for k in fields)


def _date_obj(value: str) -> datetime | None:
    if not value:
        return None
    try:
        return datetime.fromisoformat(value)
    except ValueError:
        return None


def _quality_checks(entity_type: str, payload: dict[str, Any]) -> list[tuple[str, str, str]]:
    issues: list[tuple[str, str, str]] = []
    if entity_type == "case":
        if not payload.get("full_name"):
            issues.append(("error", "Thiếu họ tên", "Ca bệnh chưa có họ tên."))
        if not payload.get("case_code"):
            issues.append(("warning", "Thiếu mã số", "Ca bệnh chưa có mã số trên hệ thống."))
        if not payload.get("main_diagnosis"):
            issues.append(("error", "Thiếu chẩn đoán", "Ca bệnh chưa có chẩn đoán chính."))
        if not payload.get("onset_date"):
            issues.append(("warning", "Thiếu ngày khởi phát", "Không xác định được ngày khởi phát."))
        onset = _date_obj(payload.get("onset_date", ""))
        report = _date_obj(payload.get("report_datetime", ""))
        if onset and report:
            delay = (report.date() - onset.date()).days
            if delay < 0:
                issues.append(("error", "Ngày báo cáo không hợp lệ", "Thời gian báo cáo trước ngày khởi phát."))
            elif delay > 2:
                issues.append(("warning", "Báo cáo muộn", f"Báo cáo sau khởi phát {delay} ngày."))
        lon, lat = payload.get("longitude"), payload.get("latitude")
        if lon is not None and not (-180 <= lon <= 180):
            issues.append(("error", "Kinh độ không hợp lệ", f"Kinh độ {lon} ngoài khoảng -180 đến 180."))
        if lat is not None and not (-90 <= lat <= 90):
            issues.append(("error", "Vĩ độ không hợp lệ", f"Vĩ độ {lat} ngoài khoảng -90 đến 90."))
    else:
        if not payload.get("disease"):
            issues.append(("error", "Thiếu tên bệnh", "Ổ dịch chưa có tên bệnh."))
        if not payload.get("location"):
            issues.append(("error", "Thiếu địa điểm", "Ổ dịch chưa có địa điểm xảy ra."))
        if payload.get("positive_count", 0) > payload.get("sample_count", 0):
            issues.append(("error", "Số mẫu không hợp lệ", "Số mẫu dương tính lớn hơn tổng số mẫu xét nghiệm."))
        if payload.get("death_count", 0) > payload.get("case_count", 0):
            issues.append(("error", "Số tử vong không hợp lệ", "Số tử vong lớn hơn số ca mắc."))
        status = normalize_key(payload.get("status"))
        if "da ket thuc" in status and not payload.get("end_date"):
            issues.append(("error", "Thiếu ngày kết thúc", "Ổ dịch đã kết thúc nhưng chưa có ngày kết thúc hoạt động."))
        if "dang hoat dong" in status and payload.get("end_date"):
            issues.append(("error", "Trạng thái không khớp", "Ổ dịch đang hoạt động nhưng đã có ngày kết thúc."))
        first = _date_obj(payload.get("first_onset_date", ""))
        last = _date_obj(payload.get("last_onset_date", ""))
        end = _date_obj(payload.get("end_date", ""))
        report = _date_obj(payload.get("report_datetime", ""))
        if first and last and last < first:
            issues.append(("error", "Ngày khởi phát không hợp lệ", "Ca cuối khởi phát trước ca đầu."))
        if first and end and end < first:
            issues.append(("error", "Ngày kết thúc không hợp lệ", "Ngày kết thúc trước ngày khởi phát ca đầu."))
        if last and end and end < last:
            issues.append(("error", "Ngày kết thúc không hợp lệ", "Ngày kết thúc trước ngày khởi phát ca cuối."))
        if first and report:
            delay = (report.date() - first.date()).days
            if delay < 0:
                issues.append(("error", "Ngày báo cáo không hợp lệ", "Ngày báo cáo trước ngày khởi phát ca đầu."))
            elif delay > 2:
                issues.append(("warning", "Báo cáo muộn", f"Báo cáo ổ dịch sau khởi phát ca đầu {delay} ngày."))
        if not payload.get("first_report_received_date"):
            issues.append(("info", "Thiếu ngày nhận báo cáo", "Chưa có ngày nhận báo cáo ổ dịch bệnh đầu tiên."))
        if not payload.get("last_onset_date"):
            issues.append(("info", "Thiếu ngày khởi phát ca cuối", "Chưa có ngày khởi phát trường hợp bệnh cuối cùng."))
    return issues


def import_excel(path: Path | str, db_path: Path | str = DB_PATH) -> ImportSummary:
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(path)
    init_db(db_path)
    entity_type, sheet_name, header_row, mapping = detect_excel(path)
    summary = ImportSummary(path.name, entity_type, sheet_name)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    reset = getattr(ws, "reset_dimensions", None)
    if callable(reset):
        reset()
    now = datetime.now().isoformat(sep=" ", timespec="seconds")
    consecutive_empty = 0
    try:
        with _connect(db_path) as conn:
            for row_no, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
                raw_payload = {field: row[idx] if idx < len(row) else None for idx, field in mapping.items()}
                payload = _normalize_payload(entity_type, raw_payload)
                if _is_empty_payload(entity_type, payload):
                    consecutive_empty += 1
                    if consecutive_empty >= 100:
                        break
                    continue
                consecutive_empty = 0
                summary.rows_read += 1
                payload["source_file"] = path.name
                payload["source_sheet"] = sheet_name
                payload["source_row"] = row_no
                payload["imported_at"] = now
                payload["raw_json"] = json.dumps({str(k): strip_text(v) for k, v in raw_payload.items()}, ensure_ascii=False)
                payload["row_hash"] = _row_hash(entity_type, payload)
                table = "cases" if entity_type == "case" else "outbreaks"
                cols = list(payload)
                placeholders = ",".join("?" for _ in cols)
                try:
                    cur = conn.execute(
                        f"INSERT INTO {table} ({','.join(cols)}) VALUES ({placeholders})",
                        [payload[c] for c in cols],
                    )
                    entity_id = cur.lastrowid
                    summary.inserted += 1
                except sqlite3.IntegrityError as exc:
                    if "row_hash" in str(exc).lower() or "unique" in str(exc).lower():
                        summary.duplicates += 1
                        continue
                    summary.skipped += 1
                    continue
                for severity, issue_type, description in _quality_checks(entity_type, payload):
                    conn.execute(
                        """INSERT INTO data_quality_issues
                           (entity_type, entity_id, source_file, source_row, severity, issue_type, description, created_at)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                        (entity_type, entity_id, path.name, row_no, severity, issue_type, description, now),
                    )
                    summary.issues += 1
            conn.execute(
                """INSERT INTO import_batches
                   (file_name, file_path, entity_type, sheet_name, rows_read, inserted, duplicates, skipped, issue_count, imported_at)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    path.name,
                    str(path),
                    entity_type,
                    sheet_name,
                    summary.rows_read,
                    summary.inserted,
                    summary.duplicates,
                    summary.skipped,
                    summary.issues,
                    now,
                ),
            )
    finally:
        wb.close()
    return summary


def dashboard_stats(db_path: Path | str = DB_PATH) -> dict[str, Any]:
    init_db(db_path)
    with _connect(db_path) as conn:
        cases = conn.execute("SELECT COUNT(*) FROM cases").fetchone()[0]
        outbreaks = conn.execute("SELECT COUNT(*) FROM outbreaks").fetchone()[0]
        active = conn.execute("SELECT COUNT(*) FROM outbreaks WHERE status LIKE '%Đang hoạt động%'").fetchone()[0]
        total_cases = conn.execute("SELECT COALESCE(SUM(case_count),0) FROM outbreaks").fetchone()[0]
        deaths = conn.execute("SELECT COALESCE(SUM(death_count),0) FROM outbreaks").fetchone()[0]
        issues = conn.execute("SELECT COUNT(*) FROM data_quality_issues WHERE severity IN ('error','warning')").fetchone()[0]
        return {
            "case_records": cases,
            "outbreak_records": outbreaks,
            "active_outbreaks": active,
            "reported_cases": total_cases,
            "deaths": deaths,
            "quality_issues": issues,
        }


def disease_summary(db_path: Path | str = DB_PATH, limit: int = 15) -> list[dict[str, Any]]:
    with _connect(db_path) as conn:
        rows = conn.execute(
            """SELECT disease, COUNT(*) AS outbreak_count,
                      COALESCE(SUM(case_count),0) AS case_count,
                      COALESCE(SUM(death_count),0) AS death_count,
                      SUM(CASE WHEN status LIKE '%Đang hoạt động%' THEN 1 ELSE 0 END) AS active_count
               FROM outbreaks
               GROUP BY disease
               ORDER BY outbreak_count DESC, disease
               LIMIT ?""",
            (limit,),
        ).fetchall()
        return [dict(r) for r in rows]


def monthly_outbreak_summary(db_path: Path | str = DB_PATH, limit: int = 18) -> list[dict[str, Any]]:
    with _connect(db_path) as conn:
        rows = conn.execute(
            """SELECT substr(first_onset_date,1,7) AS month,
                      COUNT(*) AS outbreak_count,
                      COALESCE(SUM(case_count),0) AS case_count
               FROM outbreaks
               WHERE first_onset_date <> ''
               GROUP BY substr(first_onset_date,1,7)
               ORDER BY month DESC LIMIT ?""",
            (limit,),
        ).fetchall()
        return [dict(r) for r in reversed(rows)]


def recent_active_outbreaks(db_path: Path | str = DB_PATH, limit: int = 20) -> list[dict[str, Any]]:
    with _connect(db_path) as conn:
        rows = conn.execute(
            """SELECT id, disease, location, first_onset_date, case_count, status, reporting_unit
               FROM outbreaks WHERE status LIKE '%Đang hoạt động%'
               ORDER BY first_onset_date DESC LIMIT ?""",
            (limit,),
        ).fetchall()
        return [dict(r) for r in rows]


CASE_TABLE_COLUMNS = ["id", *[db for _, db in CASE_FIELDS], "birth_year"]
OUTBREAK_TABLE_COLUMNS = ["id", *[db for _, db in OUTBREAK_FIELDS]]


def _safe_table(entity_type: str) -> tuple[str, list[str]]:
    if entity_type == "case":
        return "cases", CASE_TABLE_COLUMNS
    if entity_type == "outbreak":
        return "outbreaks", OUTBREAK_TABLE_COLUMNS
    raise ValueError("entity_type không hợp lệ")


def list_filter_values(entity_type: str, field: str, db_path: Path | str = DB_PATH) -> list[str]:
    table, columns = _safe_table(entity_type)
    allowed = set(columns) | {"record_status", "current_status", "main_diagnosis", "disease", "status", "admin_area", "reporting_unit"}
    if field not in allowed:
        raise ValueError("Trường lọc không hợp lệ")
    with _connect(db_path) as conn:
        rows = conn.execute(
            f"SELECT DISTINCT {field} FROM {table} WHERE COALESCE({field},'') <> '' ORDER BY {field}"
        ).fetchall()
        return [str(r[0]) for r in rows]


def query_records(
    entity_type: str,
    *,
    search: str = "",
    disease: str = "",
    status: str = "",
    admin_area: str = "",
    page: int = 1,
    page_size: int = 200,
    db_path: Path | str = DB_PATH,
) -> tuple[list[dict[str, Any]], int]:
    table, columns = _safe_table(entity_type)
    where: list[str] = []
    params: list[Any] = []
    if search:
        like = f"%{search}%"
        if entity_type == "case":
            where.append("(full_name LIKE ? OR case_code LIKE ? OR national_id LIKE ? OR current_address LIKE ? OR commune LIKE ?)")
            params.extend([like] * 5)
        else:
            where.append("(disease LIKE ? OR location LIKE ? OR reporting_unit LIKE ? OR admin_area LIKE ?)")
            params.extend([like] * 4)
    if disease:
        field = "main_diagnosis" if entity_type == "case" else "disease"
        where.append(f"{field} = ?")
        params.append(disease)
    if status:
        if entity_type == "case":
            where.append("(record_status = ? OR current_status = ?)")
            params.extend([status, status])
        else:
            where.append("status = ?")
            params.append(status)
    if admin_area:
        field = "commune" if entity_type == "case" else "admin_area"
        where.append(f"{field} = ?")
        params.append(admin_area)
    where_sql = " WHERE " + " AND ".join(where) if where else ""
    order = "onset_date DESC, id DESC" if entity_type == "case" else "first_onset_date DESC, id DESC"
    page = max(1, page)
    page_size = max(10, min(2000, page_size))
    with _connect(db_path) as conn:
        total = conn.execute(f"SELECT COUNT(*) FROM {table}{where_sql}", params).fetchone()[0]
        rows = conn.execute(
            f"SELECT {','.join(columns)} FROM {table}{where_sql} ORDER BY {order} LIMIT ? OFFSET ?",
            [*params, page_size, (page - 1) * page_size],
        ).fetchall()
        return [dict(r) for r in rows], total


def get_record(entity_type: str, record_id: int, db_path: Path | str = DB_PATH) -> dict[str, Any] | None:
    table, _ = _safe_table(entity_type)
    with _connect(db_path) as conn:
        row = conn.execute(f"SELECT * FROM {table} WHERE id = ?", (record_id,)).fetchone()
        return dict(row) if row else None


def save_outbreak(data: dict[str, Any], record_id: int | None = None, db_path: Path | str = DB_PATH) -> int:
    init_db(db_path)
    allowed = {db for _, db in OUTBREAK_FIELDS}
    payload = _normalize_payload("outbreak", {k: v for k, v in data.items() if k in allowed})
    payload["source_file"] = data.get("source_file") or "Nhập trực tiếp"
    payload["source_sheet"] = data.get("source_sheet") or "Ứng dụng"
    payload["source_row"] = data.get("source_row") or 0
    payload["imported_at"] = datetime.now().isoformat(sep=" ", timespec="seconds")
    payload["raw_json"] = json.dumps(data, ensure_ascii=False, default=str)
    payload["row_hash"] = _row_hash("outbreak", payload)
    with _connect(db_path) as conn:
        if record_id:
            cols = [c for c in payload if c not in {"row_hash", "imported_at"}]
            conn.execute(
                f"UPDATE outbreaks SET {','.join(f'{c}=?' for c in cols)} WHERE id=?",
                [payload[c] for c in cols] + [record_id],
            )
            conn.execute("DELETE FROM data_quality_issues WHERE entity_type='outbreak' AND entity_id=?", (record_id,))
            entity_id = record_id
        else:
            cols = list(payload)
            cur = conn.execute(
                f"INSERT INTO outbreaks ({','.join(cols)}) VALUES ({','.join('?' for _ in cols)})",
                [payload[c] for c in cols],
            )
            entity_id = int(cur.lastrowid)
        now = datetime.now().isoformat(sep=" ", timespec="seconds")
        for severity, issue_type, description in _quality_checks("outbreak", payload):
            conn.execute(
                """INSERT INTO data_quality_issues
                   (entity_type, entity_id, source_file, source_row, severity, issue_type, description, created_at)
                   VALUES ('outbreak', ?, ?, ?, ?, ?, ?, ?)""",
                (entity_id, payload["source_file"], payload["source_row"], severity, issue_type, description, now),
            )
        return entity_id


def delete_record(entity_type: str, record_id: int, db_path: Path | str = DB_PATH) -> None:
    table, _ = _safe_table(entity_type)
    create_backup(db_path)
    with _connect(db_path) as conn:
        conn.execute(f"DELETE FROM {table} WHERE id=?", (record_id,))
        conn.execute("DELETE FROM data_quality_issues WHERE entity_type=? AND entity_id=?", (entity_type, record_id))



def _match_text(value: Any) -> str:
    text = normalize_key(value)
    return re.sub(r"[^a-z0-9]+", "", text)


def _match_digits(value: Any) -> str:
    return re.sub(r"\D+", "", strip_text(value))


def _disease_match_text(value: Any) -> str:
    text = _match_text(value)
    return re.sub(r"^benh", "", text)


def _date_distance_days(left: Any, right: Any) -> int | None:
    a = _date_obj(strip_text(left))
    b = _date_obj(strip_text(right))
    if not a or not b:
        return None
    return abs((a.date() - b.date()).days)


def _weight(weights: dict[str, Any], key: str, default: int) -> int:
    try:
        return max(0, min(100, int(weights.get(key, default))))
    except (TypeError, ValueError):
        return default


def _case_pair_criteria_matches(
    a: dict[str, Any], b: dict[str, Any], criteria: CaseDuplicateCriteria,
) -> list[str]:
    """Trả về danh sách nhãn tiêu chí đã khớp giữa 2 ca bệnh (rỗng nếu không trùng).

    Không chấm điểm: CDC tự chọn tiêu chí nào coi là trùng; hai bản ghi được xem là
    trùng nếu khớp ít nhất một tiêu chí đang bật.
    """
    enabled = set(criteria.enabled)
    matches: list[str] = []
    code_a, code_b = _match_text(a.get("case_code")), _match_text(b.get("case_code"))
    id_a, id_b = _match_digits(a.get("national_id")), _match_digits(b.get("national_id"))
    phone_a, phone_b = _match_digits(a.get("phone"))[-9:], _match_digits(b.get("phone"))[-9:]
    name_a, name_b = _match_text(a.get("full_name")), _match_text(b.get("full_name"))
    commune_a, commune_b = _match_text(a.get("commune")), _match_text(b.get("commune"))

    if "case_code" in enabled and code_a and code_a == code_b:
        matches.append(CASE_CRITERIA_LABELS["case_code"])
    if "national_id" in enabled and len(id_a) >= 9 and id_a == id_b:
        matches.append(CASE_CRITERIA_LABELS["national_id"])
    if "phone" in enabled and len(phone_a) >= 7 and phone_a == phone_b:
        matches.append(CASE_CRITERIA_LABELS["phone"])
    if "name_birth_year" in enabled and name_a and name_a == name_b and a.get("birth_year") and a.get("birth_year") == b.get("birth_year"):
        matches.append(CASE_CRITERIA_LABELS["name_birth_year"])
    if "name_commune" in enabled and name_a and name_a == name_b and commune_a and commune_a == commune_b:
        matches.append(CASE_CRITERIA_LABELS["name_commune"])
    if "name_similar" in enabled and name_a and name_b and name_a != name_b:
        ratio = SequenceMatcher(None, name_a, name_b).ratio()
        if ratio * 100 >= criteria.name_similarity_percent:
            matches.append(f"{CASE_CRITERIA_LABELS['name_similar']} ({ratio:.0%})")
    if "onset_near" in enabled:
        days = _date_distance_days(a.get("onset_date"), b.get("onset_date"))
        if days is not None and days <= criteria.onset_max_days:
            matches.append(f"{CASE_CRITERIA_LABELS['onset_near']} (lệch {days} ngày)")
    return matches


def _outbreak_pair_score(a: dict[str, Any], b: dict[str, Any], weights: dict[str, Any] | None = None) -> tuple[int, list[str]]:
    weights = weights or {}
    score = 0
    reasons: list[str] = []
    disease_a, disease_b = _disease_match_text(a.get("disease")), _disease_match_text(b.get("disease"))
    location_a, location_b = _match_text(a.get("location")), _match_text(b.get("location"))
    if disease_a and disease_a == disease_b:
        score += _weight(weights, "disease", 30); reasons.append("Trùng tên bệnh")
    else:
        return 0, []
    if location_a and location_a == location_b:
        score += _weight(weights, "location_exact", 45); reasons.append("Trùng địa điểm ổ dịch")
    elif location_a and location_b:
        ratio = SequenceMatcher(None, location_a, location_b).ratio()
        if ratio >= 0.88:
            score += _weight(weights, "location_near", 35); reasons.append(f"Địa điểm gần giống {ratio:.0%}")
        elif ratio >= 0.75:
            score += round(_weight(weights, "location_near", 35) * 0.57); reasons.append(f"Địa điểm tương tự {ratio:.0%}")
    if _match_text(a.get("admin_area")) and _match_text(a.get("admin_area")) == _match_text(b.get("admin_area")):
        score += _weight(weights, "area", 10); reasons.append("Trùng địa bàn")
    days = _date_distance_days(a.get("first_onset_date"), b.get("first_onset_date"))
    if days == 0:
        score += _weight(weights, "onset_exact", 20); reasons.append("Trùng ngày khởi phát ca đầu")
    elif days is not None and days <= 7:
        score += _weight(weights, "onset_near", 15); reasons.append(f"Khởi phát ca đầu lệch {days} ngày")
    elif days is not None and days <= 14:
        score += round(_weight(weights, "onset_near", 15) * 0.53); reasons.append(f"Khởi phát ca đầu trong {days} ngày")
    if _match_text(a.get("reporting_unit")) and _match_text(a.get("reporting_unit")) == _match_text(b.get("reporting_unit")):
        score += _weight(weights, "reporting_unit", 5); reasons.append("Trùng đơn vị báo cáo")
    return min(score, 99), reasons


def find_duplicate_groups(
    entity_type: str,
    *,
    min_score: int = 65,
    max_records: int = 20000,
    rules: dict[str, Any] | None = None,
    criteria: dict[str, Any] | CaseDuplicateCriteria | None = None,
    db_path: Path | str = DB_PATH,
) -> list[dict[str, Any]]:
    """Phát hiện nhóm trùng nghiệp vụ; không tự động xóa hoặc gộp.

    Ca bệnh (``entity_type="case"``): lọc theo tiêu chí do CDC chọn (``criteria``), không
    chấm điểm — hai bản ghi trùng nếu khớp ít nhất một tiêu chí đang bật. ``min_score``/``rules``
    bị bỏ qua với ca bệnh, chỉ còn áp dụng cho ổ dịch.
    Ổ dịch (``entity_type="outbreak"``): vẫn dùng cơ chế chấm điểm/trọng số như trước.
    """
    table, _ = _safe_table(entity_type)
    if entity_type == "case":
        return _find_case_duplicate_groups(table, criteria, max_records, db_path)
    return _find_outbreak_duplicate_groups(table, min_score, rules, max_records, db_path)


def _resolve_case_criteria(criteria: dict[str, Any] | CaseDuplicateCriteria | None) -> CaseDuplicateCriteria:
    if isinstance(criteria, CaseDuplicateCriteria):
        return criteria.normalized()
    if criteria:
        return CaseDuplicateCriteria(
            enabled=list(criteria.get("enabled") or []),
            name_similarity_percent=criteria.get("name_similarity_percent", 92),
            onset_max_days=criteria.get("onset_max_days", 3),
        ).normalized()
    return load_case_criteria()


def _find_case_duplicate_groups(
    table: str, criteria: dict[str, Any] | CaseDuplicateCriteria | None, max_records: int, db_path: Path | str,
) -> list[dict[str, Any]]:
    resolved_criteria = _resolve_case_criteria(criteria)
    fields = [
        "id", "case_code", "full_name", "birth_date_raw", "birth_year", "gender",
        "national_id", "phone", "current_address", "commune", "main_diagnosis",
        "onset_date", "admission_date", "report_datetime", "reporting_unit", "source_file", "source_row",
    ]
    with _connect(db_path) as conn:
        rows = [dict(r) for r in conn.execute(
            f"SELECT {','.join(fields)} FROM {table} ORDER BY id LIMIT ?", (max_records,)
        ).fetchall()]
    if len(rows) < 2:
        return []

    # "name_similar"/"onset_near" so khớp mờ (không đòi hỏi khớp chính xác một trường nào), nên
    # không có khoá chặn tự nhiên như các tiêu chí còn lại. Khi CDC bật một trong hai, mở thêm
    # phạm vi so sánh theo từng xã — nếu không, 2 ca chỉ "tên gần giống"/"ngày khởi phát gần
    # nhau" (không trùng thêm trường nào khác) sẽ không bao giờ lọt vào cùng bucket exact-match
    # và do đó không bao giờ được so sánh, khiến 2 tiêu chí này gần như vô hiệu khi bật riêng lẻ.
    needs_commune_bucket = bool({"name_similar", "onset_near"} & set(resolved_criteria.enabled))

    buckets: dict[str, list[int]] = {}
    for index, row in enumerate(rows):
        keys: set[str] = set()
        code = _match_text(row.get("case_code"))
        national_id = _match_digits(row.get("national_id"))
        phone = _match_digits(row.get("phone"))[-9:]
        name = _match_text(row.get("full_name"))
        year = row.get("birth_year") or ""
        commune = _match_text(row.get("commune"))
        if code: keys.add("code:" + code)
        if len(national_id) >= 9: keys.add("nid:" + national_id)
        if len(phone) >= 7: keys.add("phone:" + phone)
        if name and year: keys.add(f"nameyear:{name}:{year}")
        if name and commune: keys.add(f"namearea:{name}:{commune}")
        if name: keys.add("name:" + name)
        if needs_commune_bucket and commune: keys.add("commune:" + commune)
        for key in keys:
            buckets.setdefault(key, []).append(index)

    candidate_pairs: set[tuple[int, int]] = set()
    for indexes in buckets.values():
        if len(indexes) > 250:
            indexes = indexes[:250]
        for pos, left in enumerate(indexes):
            for right in indexes[pos + 1:]:
                candidate_pairs.add((min(left, right), max(left, right)))

    edges: list[tuple[int, int, list[str]]] = []
    for left, right in candidate_pairs:
        matches = _case_pair_criteria_matches(rows[left], rows[right], resolved_criteria)
        if matches:
            edges.append((left, right, matches))
    if not edges:
        return []

    parent = list(range(len(rows)))
    def find(x: int) -> int:
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x
    def union(a: int, b: int) -> None:
        ra, rb = find(a), find(b)
        if ra != rb: parent[rb] = ra
    for left, right, _ in edges:
        union(left, right)

    groups: dict[int, set[int]] = {}
    for left, right, _ in edges:
        root = find(left)
        groups.setdefault(root, set()).update((left, right))

    definite_criteria = {CASE_CRITERIA_LABELS["case_code"], CASE_CRITERIA_LABELS["national_id"]}
    result: list[dict[str, Any]] = []
    for group_no, indexes in enumerate(sorted(groups.values(), key=lambda g: min(rows[i]["id"] for i in g)), start=1):
        group_edges = [edge for edge in edges if edge[0] in indexes and edge[1] in indexes]
        matched_criteria: list[str] = []
        for _, _, edge_matches in group_edges:
            for match in edge_matches:
                if match not in matched_criteria:
                    matched_criteria.append(match)
        records = [rows[i] for i in sorted(indexes, key=lambda i: rows[i]["id"])]
        case_codes = [str(r.get("case_code") or "") for r in records]
        summary = " / ".join(str(r.get("full_name") or r.get("case_code") or r["id"]) for r in records[:3])
        is_definite = any(any(base in match for base in definite_criteria) for match in matched_criteria)
        result.append({
            "group_id": group_no,
            "entity_type": "case",
            "confidence": "Trùng chắc chắn" if is_definite else "Nghi trùng",
            "matched_criteria": matched_criteria,
            "case_codes": case_codes,
            "record_count": len(records),
            "record_ids": [int(r["id"]) for r in records],
            "summary": summary,
            "reasons": "; ".join(matched_criteria[:8]),
            "records": records,
        })
    return sorted(result, key=lambda g: (0 if g["confidence"] == "Trùng chắc chắn" else 1, int(g["group_id"])))


def _find_outbreak_duplicate_groups(
    table: str, min_score: int, rules: dict[str, Any] | None, max_records: int, db_path: Path | str,
) -> list[dict[str, Any]]:
    entity_type = "outbreak"
    configured = load_rules()
    min_score = max(40, min(100, int(min_score if min_score is not None else configured.min_score)))
    if rules:
        weights = dict(rules.get("weights") or rules)
        definite_score = max(min_score, min(100, int(rules.get("definite_score", configured.definite_score))))
    else:
        weights = configured.weights_for(entity_type)
        definite_score = configured.definite_score
    fields = [
        "id", "disease", "location", "admin_area", "first_onset_date", "last_onset_date",
        "status", "case_count", "report_datetime", "reporting_unit", "source_file", "source_row",
    ]
    with _connect(db_path) as conn:
        rows = [dict(r) for r in conn.execute(
            f"SELECT {','.join(fields)} FROM {table} ORDER BY id LIMIT ?", (max_records,)
        ).fetchall()]
    if len(rows) < 2:
        return []

    buckets: dict[str, list[int]] = {}
    for index, row in enumerate(rows):
        keys: set[str] = set()
        disease = _disease_match_text(row.get("disease"))
        location = _match_text(row.get("location"))
        area = _match_text(row.get("admin_area"))
        if disease and location: keys.add(f"event:{disease}:{location}")
        if disease and area: keys.add(f"eventarea:{disease}:{area}")
        for key in keys:
            buckets.setdefault(key, []).append(index)

    candidate_pairs: set[tuple[int, int]] = set()
    for indexes in buckets.values():
        if len(indexes) > 250:
            indexes = indexes[:250]
        for pos, left in enumerate(indexes):
            for right in indexes[pos + 1:]:
                candidate_pairs.add((min(left, right), max(left, right)))

    edges: list[tuple[int, int, int, list[str]]] = []
    for left, right in candidate_pairs:
        score, reasons = _outbreak_pair_score(rows[left], rows[right], weights)
        if score >= min_score:
            edges.append((left, right, score, reasons))
    if not edges:
        return []

    parent = list(range(len(rows)))
    def find(x: int) -> int:
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x
    def union(a: int, b: int) -> None:
        ra, rb = find(a), find(b)
        if ra != rb: parent[rb] = ra
    for left, right, _, _ in edges:
        union(left, right)

    groups: dict[int, set[int]] = {}
    for left, right, _, _ in edges:
        root = find(left)
        groups.setdefault(root, set()).update((left, right))

    result: list[dict[str, Any]] = []
    for group_no, indexes in enumerate(sorted(groups.values(), key=lambda g: min(rows[i]["id"] for i in g)), start=1):
        group_edges = [edge for edge in edges if edge[0] in indexes and edge[1] in indexes]
        best_score = max(edge[2] for edge in group_edges)
        all_reasons: list[str] = []
        for _, _, _, edge_reasons in sorted(group_edges, key=lambda e: e[2], reverse=True):
            for reason in edge_reasons:
                if reason not in all_reasons:
                    all_reasons.append(reason)
        records = [rows[i] for i in sorted(indexes, key=lambda i: rows[i]["id"])]
        summary = " / ".join(str(r.get("location") or r["id"]) for r in records[:3])
        result.append({
            "group_id": group_no,
            "entity_type": entity_type,
            "confidence": "Trùng chắc chắn" if best_score >= definite_score else "Nghi trùng",
            "score": best_score,
            "record_count": len(records),
            "record_ids": [int(r["id"]) for r in records],
            "summary": summary,
            "reasons": "; ".join(all_reasons[:8]),
            "records": records,
        })
    return sorted(result, key=lambda g: (-int(g["score"]), int(g["group_id"])))


def _mergeable_fields(entity_type: str) -> set[str]:
    fields = {db for _, db in (CASE_FIELDS if entity_type == "case" else OUTBREAK_FIELDS)}
    fields.discard("source_stt")
    return fields


def merge_duplicate_records(
    entity_type: str,
    keep_id: int,
    remove_ids: Sequence[int],
    merged_values: dict[str, Any] | None = None,
    db_path: Path | str = DB_PATH,
    actor: str = "",
) -> dict[str, Any]:
    table, _ = _safe_table(entity_type)
    init_db(db_path)
    keep_id = int(keep_id)
    ids = sorted({int(v) for v in remove_ids if int(v) != keep_id})
    if not ids:
        raise ValueError("Chưa chọn bản ghi trùng để đưa vào thùng rác.")
    all_ids = [keep_id, *ids]
    placeholders = ",".join("?" for _ in all_ids)
    with _connect(db_path) as conn:
        records = [dict(row) for row in conn.execute(
            f"SELECT * FROM {table} WHERE id IN ({placeholders}) ORDER BY id", all_ids
        ).fetchall()]
    by_id = {int(row["id"]): row for row in records}
    if keep_id not in by_id:
        raise ValueError("Bản ghi cần giữ không tồn tại.")
    missing = [record_id for record_id in ids if record_id not in by_id]
    if missing:
        raise ValueError(f"Bản ghi cần xử lý không tồn tại: {missing}")

    allowed = _mergeable_fields(entity_type)
    requested = {k: v for k, v in (merged_values or {}).items() if k in allowed}
    normalized = _normalize_payload(entity_type, requested) if requested else {}
    normalized = {k: v for k, v in normalized.items() if k in allowed or k in {"birth_year", "admin_area"}}
    backup = create_backup(db_path)
    now = datetime.now().isoformat(sep=" ", timespec="seconds")
    with _connect(db_path) as conn:
        cur = conn.execute(
            """INSERT INTO duplicate_actions
               (entity_type, keep_id, removed_ids_json, backup_file, created_at, action_type, merged_values_json)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (entity_type, keep_id, json.dumps(ids), str(backup), now, "merge" if normalized else "remove",
             json.dumps(normalized, ensure_ascii=False, default=str)),
        )
        action_id = int(cur.lastrowid)
        for record_id in ids:
            conn.execute(
                """INSERT INTO duplicate_trash
                   (action_id, entity_type, original_id, record_json, deleted_at) VALUES (?, ?, ?, ?, ?)""",
                (action_id, entity_type, record_id, json.dumps(by_id[record_id], ensure_ascii=False, default=str), now),
            )
        remove_placeholders = ",".join("?" for _ in ids)
        conn.execute(f"DELETE FROM {table} WHERE id IN ({remove_placeholders})", ids)
        conn.execute(
            f"DELETE FROM data_quality_issues WHERE entity_type=? AND entity_id IN ({remove_placeholders})",
            [entity_type, *ids],
        )
        if normalized:
            keep = dict(by_id[keep_id])
            keep.update(normalized)
            keep["modified_datetime"] = now if entity_type == "case" else keep.get("modified_datetime")
            hash_payload = {key: keep.get(key) for key in allowed}
            new_hash = _row_hash(entity_type, hash_payload)
            conflict = conn.execute(f"SELECT id FROM {table} WHERE row_hash=? AND id<>?", (new_hash, keep_id)).fetchone()
            if conflict:
                new_hash = hashlib.sha256(f"{new_hash}:merge:{action_id}".encode()).hexdigest()
            update_values = dict(normalized)
            update_values["row_hash"] = new_hash
            if entity_type == "case":
                update_values["modified_datetime"] = now
            conn.execute(
                f"UPDATE {table} SET {','.join(f'{key}=?' for key in update_values)} WHERE id=?",
                [*update_values.values(), keep_id],
            )
            keep.update(update_values)
            conn.execute("DELETE FROM data_quality_issues WHERE entity_type=? AND entity_id=?", (entity_type, keep_id))
            for severity, issue_type, description in _quality_checks(entity_type, keep):
                conn.execute(
                    """INSERT INTO data_quality_issues
                       (entity_type, entity_id, source_file, source_row, severity, issue_type, description, created_at)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                    (entity_type, keep_id, keep.get("source_file", "Hợp nhất"), keep.get("source_row", 0),
                     severity, issue_type, description, now),
                )
    log_audit(
        "merge_duplicate_records" if normalized else "remove_duplicate_records", actor=actor,
        detail=f"entity={entity_type}; keep_id={keep_id}; removed={ids}", db_path=db_path,
    )
    return {
        "action_id": action_id, "kept_id": keep_id, "removed_ids": ids,
        "removed_count": len(ids), "merged_values": normalized, "backup_file": str(backup),
    }


def remove_duplicate_records(
    entity_type: str,
    keep_id: int,
    remove_ids: Sequence[int],
    db_path: Path | str = DB_PATH,
    actor: str = "",
) -> dict[str, Any]:
    return merge_duplicate_records(entity_type, keep_id, remove_ids, {}, db_path, actor=actor)


def list_duplicate_actions(limit: int = 200, db_path: Path | str = DB_PATH) -> list[dict[str, Any]]:
    init_db(db_path)
    with _connect(db_path) as conn:
        rows = conn.execute(
            """SELECT a.*,
                      SUM(CASE WHEN t.id IS NOT NULL AND t.restored_at IS NULL THEN 1 ELSE 0 END) AS pending_count,
                      COUNT(t.id) AS trash_count
               FROM duplicate_actions a
               LEFT JOIN duplicate_trash t ON t.action_id=a.id
               GROUP BY a.id ORDER BY a.id DESC LIMIT ?""", (max(1, int(limit)),)
        ).fetchall()
        return [dict(row) for row in rows]


def restore_duplicate_action(action_id: int, db_path: Path | str = DB_PATH, actor: str = "") -> dict[str, Any]:
    init_db(db_path)
    action_id = int(action_id)
    with _connect(db_path) as conn:
        action = conn.execute("SELECT * FROM duplicate_actions WHERE id=?", (action_id,)).fetchone()
        trash = conn.execute(
            "SELECT * FROM duplicate_trash WHERE action_id=? AND restored_at IS NULL ORDER BY id", (action_id,)
        ).fetchall()
    if not action:
        raise ValueError("Không tìm thấy thao tác lọc trùng.")
    if not trash:
        raise ValueError("Nhóm này không còn bản ghi trong thùng rác để khôi phục.")
    entity_type = str(action["entity_type"]); table, _ = _safe_table(entity_type)
    backup = create_backup(db_path)
    now = datetime.now().isoformat(sep=" ", timespec="seconds")
    restored: list[int] = []
    with _connect(db_path) as conn:
        columns = {str(row[1]) for row in conn.execute(f"PRAGMA table_info({table})").fetchall()}
        for item in trash:
            record = json.loads(item["record_json"])
            original_id = int(item["original_id"])
            requested_id = original_id if not conn.execute(f"SELECT 1 FROM {table} WHERE id=?", (original_id,)).fetchone() else None
            values = {key: value for key, value in record.items() if key in columns and key != "id"}
            row_hash = str(values.get("row_hash") or "")
            if row_hash and conn.execute(f"SELECT 1 FROM {table} WHERE row_hash=?", (row_hash,)).fetchone():
                values["row_hash"] = hashlib.sha256(f"{row_hash}:restore:{action_id}:{original_id}:{now}".encode()).hexdigest()
            if requested_id is not None:
                values = {"id": requested_id, **values}
            cols = list(values)
            cur = conn.execute(
                f"INSERT INTO {table} ({','.join(cols)}) VALUES ({','.join('?' for _ in cols)})",
                [values[col] for col in cols],
            )
            restored_id = int(requested_id if requested_id is not None else cur.lastrowid)
            restored.append(restored_id)
            for severity, issue_type, description in _quality_checks(entity_type, values):
                conn.execute(
                    """INSERT INTO data_quality_issues
                       (entity_type, entity_id, source_file, source_row, severity, issue_type, description, created_at)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                    (entity_type, restored_id, values.get("source_file", "Khôi phục"), values.get("source_row", 0),
                     severity, issue_type, description, now),
                )
            conn.execute(
                "UPDATE duplicate_trash SET restored_at=?, restored_id=? WHERE id=?",
                (now, restored_id, int(item["id"])),
            )
        conn.execute("UPDATE duplicate_actions SET restored_at=? WHERE id=?", (now, action_id))
    log_audit(
        "restore_duplicate_action", actor=actor, detail=f"action_id={action_id}; restored={restored}", db_path=db_path,
    )
    return {"action_id": action_id, "restored_ids": restored, "restored_count": len(restored), "backup_file": str(backup)}


def list_quality_issues(
    *, severity: str = "", entity_type: str = "", limit: int = 2000, db_path: Path | str = DB_PATH
) -> list[dict[str, Any]]:
    where: list[str] = []
    params: list[Any] = []
    if severity:
        where.append("severity=?")
        params.append(severity)
    if entity_type:
        where.append("entity_type=?")
        params.append(entity_type)
    sql_where = " WHERE " + " AND ".join(where) if where else ""
    with _connect(db_path) as conn:
        rows = conn.execute(
            f"""SELECT id, entity_type, entity_id, severity, issue_type, description, source_file, source_row, created_at
                FROM data_quality_issues{sql_where}
                ORDER BY CASE severity WHEN 'error' THEN 1 WHEN 'warning' THEN 2 ELSE 3 END, id DESC LIMIT ?""",
            [*params, limit],
        ).fetchall()
        return [dict(r) for r in rows]


def list_import_batches(db_path: Path | str = DB_PATH, limit: int = 50) -> list[dict[str, Any]]:
    with _connect(db_path) as conn:
        rows = conn.execute("SELECT * FROM import_batches ORDER BY id DESC LIMIT ?", (limit,)).fetchall()
        return [dict(r) for r in rows]


def _safe_path_part(value: str) -> str:
    cleaned = re.sub(r"[^\w\-]+", "_", strip_text(value), flags=re.UNICODE)
    return cleaned.strip("_") or "khong_xac_dinh"


def queue_submit(
    commune: str,
    week: str,
    file_name: str,
    file_bytes: bytes,
    *,
    source: str = "server_chinh",
    submitted_by: str = "",
    db_path: Path | str = DB_PATH,
) -> dict[str, Any]:
    """Lưu file Excel xã vừa nộp vào hàng đợi nhập liệu (chưa nhập vào CSDL chính).

    ``source`` phân biệt dữ liệu vào thẳng hàng đợi máy chủ chính (``server_chinh``) hay được
    đồng bộ về từ máy chủ phụ Google Apps Script sau khi máy chủ chính online lại (``server_phu``).
    """
    init_db(db_path)
    commune = strip_text(commune)
    week = strip_text(week)
    if not commune:
        raise ValueError("Thiếu tên xã nộp báo cáo.")
    if not week:
        raise ValueError("Thiếu tuần báo cáo (ví dụ 2026-W29).")
    if source not in QUEUE_SOURCES:
        raise ValueError("Nguồn dữ liệu không hợp lệ.")
    safe_name = Path(file_name or "du_lieu.xlsx").name
    if Path(safe_name).suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("Chỉ chấp nhận file XLSX hoặc XLSM.")
    if not file_bytes:
        raise ValueError("File rỗng hoặc thiếu nội dung.")
    if len(file_bytes) > 100 * 1024 * 1024:
        raise ValueError("File vượt quá giới hạn 100 MB.")
    now = datetime.now().isoformat(sep=" ", timespec="seconds")
    dest_dir = QUEUE_DIR / _safe_path_part(commune) / _safe_path_part(week)
    dest_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d%H%M%S")
    dest_path = dest_dir / f"{stamp}_{uuid.uuid4().hex[:8]}_{safe_name}"
    while dest_path.exists():
        dest_path = dest_dir / f"{stamp}_{uuid.uuid4().hex[:8]}_{safe_name}"
    dest_path.write_bytes(file_bytes)
    with _connect(db_path) as conn:
        cur = conn.execute(
            """INSERT INTO import_queue (commune, week, file_name, file_path, source, status, submitted_by, received_at)
               VALUES (?, ?, ?, ?, ?, 'cho_nhap', ?, ?)""",
            (commune, week, safe_name, str(dest_path), source, submitted_by, now),
        )
        queue_id = int(cur.lastrowid)
    log_audit(
        "queue_submit", actor=submitted_by or source, commune=commune,
        detail=f"week={week}; file={safe_name}; source={source}", db_path=db_path,
    )
    return {"queue_id": queue_id, "commune": commune, "week": week, "source": source, "received_at": now}


def list_import_queue(
    status: str = "", commune: str = "", limit: int = 200, offset: int = 0, db_path: Path | str = DB_PATH,
) -> list[dict[str, Any]]:
    init_db(db_path)
    where: list[str] = []
    params: list[Any] = []
    if status:
        where.append("status = ?"); params.append(status)
    if commune:
        where.append("commune = ?"); params.append(commune)
    where_sql = " WHERE " + " AND ".join(where) if where else ""
    limit = max(1, min(2000, int(limit)))
    offset = max(0, int(offset))
    with _connect(db_path) as conn:
        rows = conn.execute(
            f"SELECT * FROM import_queue{where_sql} ORDER BY commune, received_at DESC LIMIT ? OFFSET ?",
            [*params, limit, offset],
        ).fetchall()
        return [dict(r) for r in rows]


def archive_old_queue_files(older_than_days: int = 90, db_path: Path | str = DB_PATH, actor: str = "") -> dict[str, Any]:
    """Xoá file vật lý của các mục hàng đợi đã nhập từ lâu, giữ nguyên dòng CSDL để tra cứu.

    Không tự động chạy — gọi thủ công (menu/nút bảo trì) hoặc từ một tác vụ định kỳ do CDC
    thiết lập, để tránh xoá ngầm bằng chứng khi có tranh chấp.
    """
    init_db(db_path)
    older_than_days = max(1, int(older_than_days))
    cutoff = (datetime.now() - timedelta(days=older_than_days)).isoformat(sep=" ", timespec="seconds")
    with _connect(db_path) as conn:
        rows = conn.execute(
            """SELECT id, file_path FROM import_queue
               WHERE status = 'da_nhap' AND archived_at IS NULL AND imported_at IS NOT NULL AND imported_at < ?""",
            (cutoff,),
        ).fetchall()
        now = datetime.now().isoformat(sep=" ", timespec="seconds")
        archived_ids: list[int] = []
        freed_bytes = 0
        for row in rows:
            path = Path(row["file_path"])
            if path.exists():
                try:
                    freed_bytes += path.stat().st_size
                    path.unlink()
                except OSError:
                    continue
            conn.execute("UPDATE import_queue SET archived_at=? WHERE id=?", (now, row["id"]))
            archived_ids.append(int(row["id"]))
    log_audit(
        "archive_old_queue_files", actor=actor, detail=f"archived={len(archived_ids)}; older_than_days={older_than_days}",
        db_path=db_path,
    )
    return {"archived_count": len(archived_ids), "archived_ids": archived_ids, "freed_bytes": freed_bytes}


def log_audit(
    action: str, *, actor: str = "", commune: str = "", detail: str = "", db_path: Path | str = DB_PATH,
) -> None:
    """Ghi một dòng nhật ký kiểm toán: ai làm gì, ở xã nào, khi nào."""
    init_db(db_path)
    now = datetime.now().isoformat(sep=" ", timespec="seconds")
    with _connect(db_path) as conn:
        conn.execute(
            "INSERT INTO audit_log (created_at, actor, action, commune, detail) VALUES (?, ?, ?, ?, ?)",
            (now, actor or "he_thong", action, commune or "", detail or ""),
        )


def list_audit_log(
    limit: int = 200, action: str = "", commune: str = "", db_path: Path | str = DB_PATH,
) -> list[dict[str, Any]]:
    init_db(db_path)
    where: list[str] = []
    params: list[Any] = []
    if action:
        where.append("action = ?"); params.append(action)
    if commune:
        where.append("commune = ?"); params.append(commune)
    where_sql = " WHERE " + " AND ".join(where) if where else ""
    limit = max(1, min(2000, int(limit)))
    with _connect(db_path) as conn:
        rows = conn.execute(
            f"SELECT * FROM audit_log{where_sql} ORDER BY id DESC LIMIT ?", [*params, limit]
        ).fetchall()
        return [dict(r) for r in rows]


def _hash_password(password: str, salt: bytes | None = None, iterations: int = 200_000) -> str:
    salt = salt or os.urandom(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, iterations)
    return f"{iterations}${salt.hex()}${digest.hex()}"


def _verify_password(password: str, stored: str) -> bool:
    try:
        iterations_str, salt_hex, digest_hex = stored.split("$")
        salt = bytes.fromhex(salt_hex)
        expected = bytes.fromhex(digest_hex)
    except (ValueError, AttributeError):
        return False
    actual = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, int(iterations_str))
    return hmac.compare_digest(actual, expected)


def has_commune_accounts(db_path: Path | str = DB_PATH) -> bool:
    init_db(db_path)
    with _connect(db_path) as conn:
        row = conn.execute("SELECT 1 FROM commune_accounts WHERE active=1 LIMIT 1").fetchone()
        return row is not None


def create_commune_account(
    commune: str, username: str, password: str, display_name: str = "", db_path: Path | str = DB_PATH,
    actor: str = "",
) -> dict[str, Any]:
    init_db(db_path)
    commune = strip_text(commune)
    username = strip_text(username).lower()
    if not commune:
        raise ValueError("Thiếu tên xã.")
    if not username:
        raise ValueError("Thiếu tên đăng nhập.")
    if len(password or "") < 8:
        raise ValueError("Mật khẩu phải có ít nhất 8 ký tự.")
    now = datetime.now().isoformat(sep=" ", timespec="seconds")
    with _connect(db_path) as conn:
        try:
            cur = conn.execute(
                """INSERT INTO commune_accounts (commune, username, password_hash, display_name, active, created_at)
                   VALUES (?, ?, ?, ?, 1, ?)""",
                (commune, username, _hash_password(password), display_name or commune, now),
            )
        except sqlite3.IntegrityError as exc:
            raise ValueError("Xã hoặc tên đăng nhập này đã có tài khoản.") from exc
        account_id = int(cur.lastrowid)
    log_audit("create_commune_account", actor=actor, commune=commune, detail=f"username={username}", db_path=db_path)
    return {"id": account_id, "commune": commune, "username": username}


def list_commune_accounts(db_path: Path | str = DB_PATH) -> list[dict[str, Any]]:
    init_db(db_path)
    with _connect(db_path) as conn:
        rows = conn.execute(
            "SELECT id, commune, username, display_name, active, created_at, last_login_at "
            "FROM commune_accounts ORDER BY commune"
        ).fetchall()
        return [dict(r) for r in rows]


def set_commune_account_active(account_id: int, active: bool, db_path: Path | str = DB_PATH, actor: str = "") -> None:
    init_db(db_path)
    with _connect(db_path) as conn:
        row = conn.execute("SELECT commune FROM commune_accounts WHERE id=?", (int(account_id),)).fetchone()
        if not row:
            raise ValueError("Không tìm thấy tài khoản.")
        conn.execute("UPDATE commune_accounts SET active=? WHERE id=?", (1 if active else 0, int(account_id)))
    log_audit(
        "enable_commune_account" if active else "disable_commune_account",
        actor=actor, commune=row["commune"], db_path=db_path,
    )


def reset_commune_account_password(
    account_id: int, new_password: str, db_path: Path | str = DB_PATH, actor: str = "",
) -> None:
    init_db(db_path)
    if len(new_password or "") < 8:
        raise ValueError("Mật khẩu phải có ít nhất 8 ký tự.")
    with _connect(db_path) as conn:
        row = conn.execute("SELECT commune FROM commune_accounts WHERE id=?", (int(account_id),)).fetchone()
        if not row:
            raise ValueError("Không tìm thấy tài khoản.")
        conn.execute(
            "UPDATE commune_accounts SET password_hash=? WHERE id=?", (_hash_password(new_password), int(account_id))
        )
    log_audit("reset_commune_account_password", actor=actor, commune=row["commune"], db_path=db_path)


def verify_commune_account(username: str, password: str, db_path: Path | str = DB_PATH) -> dict[str, Any] | None:
    """Kiểm tra tên đăng nhập/mật khẩu tài khoản xã; ghi nhật ký cả khi thành công lẫn thất bại."""
    init_db(db_path)
    username_norm = strip_text(username).lower()
    with _connect(db_path) as conn:
        row = conn.execute(
            "SELECT * FROM commune_accounts WHERE username=? AND active=1", (username_norm,)
        ).fetchone()
        if not row or not _verify_password(password or "", row["password_hash"]):
            log_audit("login_failed", actor=username or "khong_ro", db_path=db_path)
            return None
        now = datetime.now().isoformat(sep=" ", timespec="seconds")
        conn.execute("UPDATE commune_accounts SET last_login_at=? WHERE id=?", (now, row["id"]))
    log_audit("login", actor=row["username"], commune=row["commune"], db_path=db_path)
    return {
        "id": int(row["id"]), "commune": row["commune"], "username": row["username"],
        "display_name": row["display_name"] or row["commune"],
    }


def issue_commune_token(
    account_id: int, commune: str, username: str, secret: str, ttl_seconds: int = 8 * 3600,
) -> str:
    """Token đăng nhập tự chứa (stateless), ký HMAC — không cần lưu phiên trên máy chủ."""
    if not secret:
        raise ValueError("Máy chủ chưa có khóa ký phiên đăng nhập (web_token_secret).")
    expires = int(time.time()) + max(1, int(ttl_seconds))
    payload = f"{account_id}:{commune}:{username}:{expires}"
    signature = hmac.new(secret.encode("utf-8"), payload.encode("utf-8"), hashlib.sha256).hexdigest()
    return base64.urlsafe_b64encode(f"{payload}:{signature}".encode("utf-8")).decode("ascii")


def verify_commune_token(token: str, secret: str) -> dict[str, Any] | None:
    if not token or not secret:
        return None
    try:
        raw = base64.urlsafe_b64decode(token.encode("ascii")).decode("utf-8")
        account_id_str, commune, username, expires_str, signature = raw.split(":", 4)
        payload = f"{account_id_str}:{commune}:{username}:{expires_str}"
        expected = hmac.new(secret.encode("utf-8"), payload.encode("utf-8"), hashlib.sha256).hexdigest()
        if not hmac.compare_digest(signature, expected):
            return None
        if int(expires_str) < int(time.time()):
            return None
        return {"account_id": int(account_id_str), "commune": commune, "username": username}
    except Exception:
        return None


def has_any_data(db_path: Path | str = DB_PATH) -> bool:
    """Máy chủ đã có dữ liệu thật (ca bệnh, ổ dịch, hoặc tài khoản) hay còn hoàn toàn trống —
    dùng làm lớp bảo vệ trước khi nhận 1 bản sao lưu đầy đủ từ máy chủ khác (di chuyển máy chủ,
    xem lan_server._handle_receive_full_backup), tránh ghi đè nhầm dữ liệu đang hoạt động."""
    init_db(db_path)
    with _connect(db_path) as conn:
        for table in ("cases", "outbreaks", "commune_accounts", "cdc_accounts"):
            if conn.execute(f"SELECT 1 FROM {table} LIMIT 1").fetchone():
                return True
    return False


def has_cdc_accounts(db_path: Path | str = DB_PATH) -> bool:
    """Đã có ít nhất 1 tài khoản quản trị viên riêng chưa — dùng để quyết định máy trạm có bắt
    đăng nhập cá nhân hay còn dùng mật khẩu máy chủ dùng chung (giai đoạn chuyển tiếp)."""
    init_db(db_path)
    with _connect(db_path) as conn:
        row = conn.execute("SELECT 1 FROM cdc_accounts WHERE active=1 LIMIT 1").fetchone()
        return row is not None


def create_cdc_account(
    username: str, password: str, display_name: str = "", db_path: Path | str = DB_PATH, actor: str = "",
) -> dict[str, Any]:
    init_db(db_path)
    username = strip_text(username).lower()
    if not username:
        raise ValueError("Thiếu tên đăng nhập.")
    if len(password or "") < 8:
        raise ValueError("Mật khẩu phải có ít nhất 8 ký tự.")
    now = datetime.now().isoformat(sep=" ", timespec="seconds")
    with _connect(db_path) as conn:
        try:
            cur = conn.execute(
                """INSERT INTO cdc_accounts (username, password_hash, display_name, active, created_at)
                   VALUES (?, ?, ?, 1, ?)""",
                (username, _hash_password(password), display_name or username, now),
            )
        except sqlite3.IntegrityError as exc:
            raise ValueError("Tên đăng nhập này đã có tài khoản.") from exc
        account_id = int(cur.lastrowid)
    log_audit("create_cdc_account", actor=actor, detail=f"username={username}", db_path=db_path)
    return {"id": account_id, "username": username}


def list_cdc_accounts(db_path: Path | str = DB_PATH) -> list[dict[str, Any]]:
    init_db(db_path)
    with _connect(db_path) as conn:
        rows = conn.execute(
            "SELECT id, username, display_name, active, created_at, last_login_at "
            "FROM cdc_accounts ORDER BY username"
        ).fetchall()
        return [dict(r) for r in rows]


def set_cdc_account_active(account_id: int, active: bool, db_path: Path | str = DB_PATH, actor: str = "") -> None:
    init_db(db_path)
    with _connect(db_path) as conn:
        row = conn.execute("SELECT username FROM cdc_accounts WHERE id=?", (int(account_id),)).fetchone()
        if not row:
            raise ValueError("Không tìm thấy tài khoản.")
        conn.execute("UPDATE cdc_accounts SET active=? WHERE id=?", (1 if active else 0, int(account_id)))
    log_audit(
        "enable_cdc_account" if active else "disable_cdc_account",
        actor=actor, detail=f"username={row['username']}", db_path=db_path,
    )


def reset_cdc_account_password(
    account_id: int, new_password: str, db_path: Path | str = DB_PATH, actor: str = "",
) -> None:
    init_db(db_path)
    if len(new_password or "") < 8:
        raise ValueError("Mật khẩu phải có ít nhất 8 ký tự.")
    with _connect(db_path) as conn:
        row = conn.execute("SELECT username FROM cdc_accounts WHERE id=?", (int(account_id),)).fetchone()
        if not row:
            raise ValueError("Không tìm thấy tài khoản.")
        conn.execute(
            "UPDATE cdc_accounts SET password_hash=? WHERE id=?", (_hash_password(new_password), int(account_id))
        )
    log_audit("reset_cdc_account_password", actor=actor, detail=f"username={row['username']}", db_path=db_path)


def verify_cdc_account(username: str, password: str, db_path: Path | str = DB_PATH) -> dict[str, Any] | None:
    """Kiểm tra tên đăng nhập/mật khẩu tài khoản quản trị viên; ghi nhật ký cả khi thành công
    lẫn thất bại — dùng cho máy trạm quản trị đăng nhập cá nhân (POST /cdc/login)."""
    init_db(db_path)
    username_norm = strip_text(username).lower()
    with _connect(db_path) as conn:
        row = conn.execute(
            "SELECT * FROM cdc_accounts WHERE username=? AND active=1", (username_norm,)
        ).fetchone()
        if not row or not _verify_password(password or "", row["password_hash"]):
            log_audit("login_failed", actor=username or "khong_ro", db_path=db_path)
            return None
        now = datetime.now().isoformat(sep=" ", timespec="seconds")
        conn.execute("UPDATE cdc_accounts SET last_login_at=? WHERE id=?", (now, row["id"]))
    log_audit("login", actor=row["username"], db_path=db_path)
    return {
        "id": int(row["id"]), "username": row["username"], "display_name": row["display_name"] or row["username"],
    }


def issue_admin_token(account_id: int, username: str, secret: str, ttl_seconds: int = 8 * 3600) -> str:
    """Token đăng nhập quản trị viên tự chứa (stateless), ký HMAC — cùng cơ chế
    issue_commune_token nhưng không gắn với xã nào (payload không có trường commune)."""
    if not secret:
        raise ValueError("Máy chủ chưa có khóa ký phiên đăng nhập (web_token_secret).")
    expires = int(time.time()) + max(1, int(ttl_seconds))
    payload = f"{account_id}:{username}:{expires}"
    signature = hmac.new(secret.encode("utf-8"), payload.encode("utf-8"), hashlib.sha256).hexdigest()
    return base64.urlsafe_b64encode(f"{payload}:{signature}".encode("utf-8")).decode("ascii")


def verify_admin_token(token: str, secret: str) -> dict[str, Any] | None:
    if not token or not secret:
        return None
    try:
        raw = base64.urlsafe_b64decode(token.encode("ascii")).decode("utf-8")
        account_id_str, username, expires_str, signature = raw.split(":", 3)
        payload = f"{account_id_str}:{username}:{expires_str}"
        expected = hmac.new(secret.encode("utf-8"), payload.encode("utf-8"), hashlib.sha256).hexdigest()
        if not hmac.compare_digest(signature, expected):
            return None
        if int(expires_str) < int(time.time()):
            return None
        return {"account_id": int(account_id_str), "username": username}
    except Exception:
        return None


def import_queue_item(queue_id: int, db_path: Path | str = DB_PATH, actor: str = "") -> dict[str, Any]:
    """Nhập một file đang chờ trong hàng đợi vào CSDL chính bằng import_excel hiện có.

    Dùng một câu UPDATE nguyên tử để "giữ chỗ" mục hàng đợi (chuyển sang trạng thái tạm
    ``dang_nhap``) trước khi nhập — tránh 2 yêu cầu đồng thời (2 người CDC cùng bấm nhập, hoặc
    bấm đúp) cùng đọc thấy ``cho_nhap`` rồi cùng chạy ``import_excel`` song song trên một file.
    """
    init_db(db_path)
    queue_id = int(queue_id)
    with _connect(db_path) as conn:
        row = conn.execute("SELECT * FROM import_queue WHERE id=?", (queue_id,)).fetchone()
        if not row:
            raise ValueError("Không tìm thấy mục trong hàng đợi.")
        item = dict(row)
        claimed = conn.execute(
            "UPDATE import_queue SET status='dang_nhap' WHERE id=? AND status='cho_nhap'", (queue_id,)
        )
        if claimed.rowcount == 0:
            if item["status"] == "da_nhap":
                raise ValueError("Mục này đã được nhập vào CSDL trước đó.")
            raise ValueError("Mục này đang được xử lý hoặc đã nhập bởi một thao tác khác.")
    file_path = Path(item["file_path"])
    now = datetime.now().isoformat(sep=" ", timespec="seconds")
    if not file_path.exists():
        with _connect(db_path) as conn:
            conn.execute(
                "UPDATE import_queue SET status='loi', error_message=? WHERE id=?",
                ("Không tìm thấy file đã lưu trên máy chủ.", queue_id),
            )
        raise ValueError("Không tìm thấy file đã lưu trên máy chủ.")
    try:
        summary = import_excel(file_path, db_path)
    except Exception as exc:
        with _connect(db_path) as conn:
            conn.execute("UPDATE import_queue SET status='loi', error_message=? WHERE id=?", (str(exc), queue_id))
        raise
    with _connect(db_path) as conn:
        batch = conn.execute(
            "SELECT id FROM import_batches WHERE file_name=? ORDER BY id DESC LIMIT 1", (file_path.name,)
        ).fetchone()
        conn.execute(
            "UPDATE import_queue SET status='da_nhap', imported_at=?, import_batch_id=?, error_message=NULL WHERE id=?",
            (now, batch[0] if batch else None, queue_id),
        )
    log_audit(
        "import_queue_item", actor=actor, commune=item.get("commune", ""),
        detail=f"queue_id={queue_id}; file={item['file_name']}; inserted={summary.inserted}", db_path=db_path,
    )
    return {
        "queue_id": queue_id, "file_name": item["file_name"], "rows_read": summary.rows_read,
        "inserted": summary.inserted, "duplicates": summary.duplicates, "skipped": summary.skipped,
        "issues": summary.issues, "summary_text": summary.as_text(),
    }


def execute_select(sql: str, db_path: Path | str = DB_PATH, max_rows: int = 5000) -> tuple[list[str], list[list[Any]]]:
    cleaned = sql.strip().rstrip(";")
    if not re.match(r"^(SELECT|WITH)\b", cleaned, flags=re.I):
        raise ValueError("Chỉ cho phép câu lệnh SELECT hoặc WITH ... SELECT.")
    forbidden = re.search(r"\b(INSERT|UPDATE|DELETE|DROP|ALTER|ATTACH|DETACH|PRAGMA|REPLACE|VACUUM|CREATE)\b", cleaned, flags=re.I)
    if forbidden:
        raise ValueError(f"Không cho phép từ khóa {forbidden.group(1).upper()} trong truy vấn.")
    with _connect(db_path) as conn:
        cur = conn.execute(cleaned)
        columns = [d[0] for d in cur.description or []]
        rows = [list(r) for r in cur.fetchmany(max_rows)]
        return columns, rows


def export_rows(path: Path | str, columns: Sequence[str], rows: Iterable[Sequence[Any]]) -> None:
    path = Path(path)
    rows = list(rows)
    if path.suffix.lower() == ".csv":
        with path.open("w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(columns)
            writer.writerows(rows)
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Dữ liệu"
    ws.append(list(columns))
    for row in rows:
        ws.append(list(row))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    for col in ws.columns:
        values = [str(c.value or "") for c in col[:200]]
        width = min(max(max((len(v) for v in values), default=8) + 2, 10), 45)
        ws.column_dimensions[col[0].column_letter].width = width
    wb.save(path)


def export_filtered_records(
    path: Path | str,
    entity_type: str,
    *,
    search: str = "",
    disease: str = "",
    status: str = "",
    admin_area: str = "",
    db_path: Path | str = DB_PATH,
) -> int:
    table, _ = _safe_table(entity_type)
    where: list[str] = []
    params: list[Any] = []
    if search:
        like = f"%{search}%"
        if entity_type == "case":
            where.append("(full_name LIKE ? OR case_code LIKE ? OR national_id LIKE ? OR current_address LIKE ? OR commune LIKE ?)")
            params.extend([like] * 5)
        else:
            where.append("(disease LIKE ? OR location LIKE ? OR reporting_unit LIKE ? OR admin_area LIKE ?)")
            params.extend([like] * 4)
    if disease:
        field = "main_diagnosis" if entity_type == "case" else "disease"
        where.append(f"{field} = ?")
        params.append(disease)
    if status:
        if entity_type == "case":
            where.append("(record_status = ? OR current_status = ?)")
            params.extend([status, status])
        else:
            where.append("status = ?")
            params.append(status)
    if admin_area:
        field = "commune" if entity_type == "case" else "admin_area"
        where.append(f"{field} = ?")
        params.append(admin_area)
    where_sql = " WHERE " + " AND ".join(where) if where else ""
    order = "onset_date DESC, id DESC" if entity_type == "case" else "first_onset_date DESC, id DESC"
    hidden = {"row_hash", "raw_json"}
    with _connect(db_path) as conn:
        total = conn.execute(f"SELECT COUNT(*) FROM {table}{where_sql}", params).fetchone()[0]
        if total == 0:
            raise ValueError("Không có dữ liệu phù hợp để xuất.")
        if total > 50_000:
            raise ValueError("Bộ lọc có trên 50.000 dòng. Hãy thu hẹp bộ lọc trước khi xuất.")
        db_columns = [r[1] for r in conn.execute(f"PRAGMA table_info({table})").fetchall() if r[1] not in hidden]
        values = conn.execute(
            f"SELECT {','.join(db_columns)} FROM {table}{where_sql} ORDER BY {order}", params
        ).fetchall()
    labels = CASE_LABELS if entity_type == "case" else OUTBREAK_LABELS
    extra_labels = {
        "id": "ID",
        "birth_year": "Năm sinh",
        "admin_area": "Địa bàn chuẩn hóa",
        "source_file": "File nguồn",
        "source_sheet": "Sheet nguồn",
        "source_row": "Dòng nguồn",
        "imported_at": "Thời điểm nhập",
    }
    out_headers = [labels.get(c, extra_labels.get(c, c)) for c in db_columns]
    export_rows(path, out_headers, [[row[c] for c in db_columns] for row in values])
    return total


_INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")


def _safe_sheet_name(name: str, used: set[str]) -> str:
    cleaned = _INVALID_SHEET_CHARS.sub(" ", (name or "").strip()) or "Chua_xac_dinh"
    cleaned = cleaned[:31] or "Sheet"
    base = cleaned
    counter = 1
    while cleaned in used:
        counter += 1
        suffix = f" ({counter})"
        cleaned = base[: 31 - len(suffix)] + suffix
    used.add(cleaned)
    return cleaned


def export_cases_by_commune(
    path: Path | str,
    *,
    criteria: dict[str, Any] | CaseDuplicateCriteria | None = None,
    db_path: Path | str = DB_PATH,
    actor: str = "",
) -> dict[str, Any]:
    """Xuất toàn bộ ca bệnh thành workbook Excel chia theo xã, mỗi xã một sheet.

    Dùng để gửi các xã tự lọc trùng trên phần mềm của Bộ Y tế. Khi một nhóm ca trùng có bản ghi
    thuộc nhiều xã khác nhau, cả nhóm được xếp vào sheet của xã có ``admission_date`` (ngày vào
    viện) gần hiện tại nhất — phản ánh nơi bệnh nhân đang được xử lý thực tế; nếu không phân
    định được thì lần lượt so ``onset_date`` rồi ``report_datetime``, cuối cùng giữ xã của bản
    ghi có id nhỏ nhất và đánh dấu "cần xác nhận" trên sheet Tong_hop.
    """
    path = Path(path)
    init_db(db_path)
    resolved_criteria = _resolve_case_criteria(criteria)
    with _connect(db_path) as conn:
        total_cases = int(conn.execute("SELECT COUNT(*) FROM cases").fetchone()[0])
    # Dò trùng phải phủ hết dữ liệu sẽ xuất bên dưới (không LIMIT) — nếu không, các ca có id
    # lớn hơn ngưỡng mặc định của find_duplicate_groups vẫn được xuất nhưng chưa được dò trùng.
    groups = _find_case_duplicate_groups("cases", resolved_criteria, max(total_cases, 1), db_path)

    hidden = {"row_hash", "raw_json"}
    with _connect(db_path) as conn:
        db_columns = [r[1] for r in conn.execute("PRAGMA table_info(cases)").fetchall() if r[1] not in hidden]
        rows = [dict(zip(db_columns, r)) for r in conn.execute(
            f"SELECT {','.join(db_columns)} FROM cases ORDER BY commune, id"
        ).fetchall()]
    if not rows:
        raise ValueError("Không có ca bệnh để xuất.")

    def _record_sort_key(record: dict[str, Any]) -> tuple[datetime, datetime, datetime, int]:
        admission = _date_obj(strip_text(record.get("admission_date") or "")) or datetime.min
        onset = _date_obj(strip_text(record.get("onset_date") or "")) or datetime.min
        report = _date_obj(strip_text(record.get("report_datetime") or "")) or datetime.min
        return (admission, onset, report, -int(record["id"]))

    commune_override: dict[int, str] = {}
    case_group_info: dict[int, dict[str, Any]] = {}
    group_summary: list[dict[str, Any]] = []
    for group in groups:
        communes = {str(r.get("commune") or "").strip() for r in group["records"]}
        communes.discard("")
        cross_commune = len(communes) > 1
        resolved_commune = ""
        needs_confirmation = False
        if cross_commune:
            winner = max(group["records"], key=_record_sort_key)
            resolved_commune = str(winner.get("commune") or "").strip()
            needs_confirmation = not any(
                _date_obj(strip_text(r.get("admission_date") or "")) for r in group["records"]
            )
            for record in group["records"]:
                commune_override[int(record["id"])] = resolved_commune
        for record in group["records"]:
            case_group_info[int(record["id"])] = {"group_id": group["group_id"], "reasons": group["reasons"]}
        group_summary.append({
            "group_id": group["group_id"],
            "confidence": group["confidence"],
            "case_codes": ", ".join(code for code in group.get("case_codes") or [] if code),
            "communes_goc": ", ".join(sorted(c for c in communes if c)) or "(cùng xã)",
            "xa_duoc_chon": resolved_commune if cross_commune else "(cùng xã, không đổi)",
            "tieu_chi_khop": group["reasons"],
            "can_xac_nhan": "Có" if needs_confirmation else "",
        })

    by_commune: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        commune = commune_override.get(int(row["id"])) or str(row.get("commune") or "").strip() or "Chưa xác định xã"
        by_commune.setdefault(commune, []).append(row)

    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Tong_hop"
    summary_ws.append(["Nhóm", "Mức", "Mã ca bệnh liên quan", "Xã gốc các bản ghi", "Xã được chọn", "Tiêu chí khớp", "Cần xác nhận"])
    for item in group_summary:
        summary_ws.append([
            item["group_id"], item["confidence"], item["case_codes"], item["communes_goc"],
            item["xa_duoc_chon"], item["tieu_chi_khop"], item["can_xac_nhan"],
        ])
    if not group_summary:
        summary_ws.append(["Không phát hiện ca trùng theo tiêu chí hiện tại.", "", "", "", "", "", ""])
    for cell in summary_ws[1]:
        cell.font = cell.font.copy(bold=True)
    summary_ws.freeze_panes = "A2"

    extra_labels = {
        "id": "ID", "birth_year": "Năm sinh", "source_file": "File nguồn",
        "source_sheet": "Sheet nguồn", "source_row": "Dòng nguồn", "imported_at": "Thời điểm nhập",
    }
    out_headers = [CASE_LABELS.get(c, extra_labels.get(c, c)) for c in db_columns] + ["Nhóm trùng", "Ghi chú lọc trùng"]

    used_sheet_names: set[str] = {"Tong_hop"}
    for commune in sorted(by_commune):
        ws = wb.create_sheet(_safe_sheet_name(commune, used_sheet_names))
        ws.append(out_headers)
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)
        for row in sorted(by_commune[commune], key=lambda r: r["id"]):
            info = case_group_info.get(int(row["id"]))
            values = [row.get(c) for c in db_columns]
            values.append(info["group_id"] if info else "")
            values.append(info["reasons"] if info else "")
            ws.append(values)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for col in ws.columns:
            col_values = [str(c.value or "") for c in col[:200]]
            width = min(max(max((len(v) for v in col_values), default=8) + 2, 10), 45)
            ws.column_dimensions[col[0].column_letter].width = width

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    log_audit(
        "export_cases_by_commune", actor=actor,
        detail=f"cases={len(rows)}; communes={len(by_commune)}; groups={len(groups)}", db_path=db_path,
    )
    return {
        "path": str(path),
        "case_count": len(rows),
        "commune_count": len(by_commune),
        "duplicate_group_count": len(groups),
        "cross_commune_group_count": sum(1 for g in group_summary if g["xa_duoc_chon"] not in ("", "(cùng xã, không đổi)")),
    }


def _backup_policy_for_core():
    policy = backup_manager.load_policy()
    if not policy.destination:
        policy.destination = str(BACKUP_DIR)
    return policy


def create_backup(db_path: Path | str = DB_PATH) -> Path:
    init_db(db_path)
    return backup_manager.create_backup(db_path, kind="manual", policy=_backup_policy_for_core())


def list_backups() -> list[dict[str, Any]]:
    return backup_manager.list_backups(_backup_policy_for_core())


def verify_backup(path: Path | str) -> dict[str, Any]:
    return backup_manager.verify_backup(path)


def restore_backup(path: Path | str, db_path: Path | str = DB_PATH) -> dict[str, Any]:
    return backup_manager.restore_backup(path, db_path)


def open_folder(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)
    if sys.platform.startswith("win"):
        os.startfile(path)  # type: ignore[attr-defined]
    elif sys.platform == "darwin":
        os.system(f'open "{path}"')
    else:
        os.system(f'xdg-open "{path}" >/dev/null 2>&1 &')


init_db()
