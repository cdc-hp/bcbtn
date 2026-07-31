from pathlib import Path
import tempfile

from openpyxl import Workbook

import core


def make_case_file(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Disease Cases'
    ws.append([label for label, _ in core.CASE_FIELDS])
    ws.append([
        1, 'CA-001', 'Nguyễn Văn A', '01/01/1990', 'Nhân viên', '', '', '', '', 'Kinh', 'Nam',
        '012345678901', '0900000000', 'Hải Phòng', 'Thành phố Hải Phòng', 'Phường Gia Viên', '',
        106.68, 20.86, 'Bệnh sốt xuất huyết Dengue', 'Nhẹ', '', 'Không rõ', 0, 'Xác định', 'Có',
        '10/07/2026', 'CDC Hải Phòng', 'PCR', 'Dương tính', '', 'Đang điều trị', '', '08/07/2026',
        '09/07/2026', '', '', 'Cán bộ A', '0900000001', 'a@example.com', 'Trạm y tế Phường Gia Viên',
        'Thành phố Hải Phòng', 'Bệnh viện A', 'Đã duyệt', '09/07/2026 08:00', '', '', '09/07/2026 09:00'
    ])
    wb.save(path)


def make_outbreak_file(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Danh sách ổ dịch'
    ws.append(['DANH SÁCH Ổ DỊCH'])
    ws.append([])
    ws.append([label for label, _ in core.OUTBREAK_FIELDS])
    ws.append([
        1, 'Bệnh sốt xuất huyết Dengue', 'Tổ 1 - Phường Gia Viên - Thành phố Hải Phòng',
        '08/07/2026', '', 'Đang hoạt động', 2, 0, 2, 2, '09/07/2026 08:00',
        'Trạm y tế Phường Gia Viên', 'Thành phố Hải Phòng', '09/07/2026', '09/07/2026'
    ])
    wb.save(path)


def test_detect_and_import_files():
    with tempfile.TemporaryDirectory() as tmp:
        case_file = Path(tmp) / 'cases.xlsx'
        outbreak_file = Path(tmp) / 'outbreaks.xlsx'
        db = Path(tmp) / 'test.db'
        make_case_file(case_file)
        make_outbreak_file(outbreak_file)
        assert core.detect_excel(case_file)[0] == 'case'
        assert core.detect_excel(outbreak_file)[0] == 'outbreak'
        assert core.import_excel(case_file, db).inserted == 1
        assert core.import_excel(outbreak_file, db).inserted == 1
        stats = core.dashboard_stats(db)
        assert stats['case_records'] == 1
        assert stats['outbreak_records'] == 1


def test_deduplicate():
    with tempfile.TemporaryDirectory() as tmp:
        outbreak_file = Path(tmp) / 'outbreaks.xlsx'
        db = Path(tmp) / 'test.db'
        make_outbreak_file(outbreak_file)
        first = core.import_excel(outbreak_file, db)
        second = core.import_excel(outbreak_file, db)
        assert first.inserted == 1
        assert second.inserted == 0
        assert second.duplicates == 1


def test_query_and_select():
    with tempfile.TemporaryDirectory() as tmp:
        outbreak_file = Path(tmp) / 'outbreaks.xlsx'
        db = Path(tmp) / 'test.db'
        make_outbreak_file(outbreak_file)
        core.import_excel(outbreak_file, db)
        rows, total = core.query_records('outbreak', disease='Bệnh sốt xuất huyết Dengue', db_path=db)
        assert total == 1
        assert rows
        cols, values = core.execute_select('SELECT COUNT(*) AS n FROM outbreaks', db_path=db)
        assert cols == ['n']
        assert values[0][0] == 1


def test_read_only_sql_guard():
    with tempfile.TemporaryDirectory() as tmp:
        db = Path(tmp) / 'test.db'
        core.init_db(db)
        try:
            core.execute_select('DELETE FROM outbreaks', db_path=db)
            raise AssertionError('DELETE should be rejected')
        except ValueError:
            pass


def corrupt_sheet_dimension(path: Path) -> None:
    """Giả lập file XLSX từ hệ thống xuất sai dimension=A1 dù có đủ cột/dòng."""
    import zipfile
    temp = path.with_suffix('.tmp.xlsx')
    with zipfile.ZipFile(path, 'r') as src, zipfile.ZipFile(temp, 'w', zipfile.ZIP_DEFLATED) as dst:
        for item in src.infolist():
            data = src.read(item.filename)
            if item.filename == 'xl/worksheets/sheet1.xml':
                text = data.decode('utf-8')
                import re
                text = re.sub(r'<dimension ref="[^"]+"', '<dimension ref="A1"', text, count=1)
                data = text.encode('utf-8')
            dst.writestr(item, data)
    temp.replace(path)


def test_detect_case_with_bad_dimension_and_header_variants():
    with tempfile.TemporaryDirectory() as tmp:
        case_file = Path(tmp) / 'Danh_sach_ca_benh (1).xlsx'
        db = Path(tmp) / 'test.db'
        wb = Workbook()
        ws = wb.active
        ws.title = 'Disease Cases'
        headers = [label for label, _ in core.CASE_FIELDS]
        headers[1] = 'Mã ca bệnh *'
        headers[2] = 'HỌ VÀ TÊN\n'
        headers[11] = 'CCCD/CMND'
        headers[12] = 'Số điện thoại'
        headers[44] = 'Ngày giờ báo cáo'
        ws.append(headers)
        ws.append([
            1, 'CA-002', 'Trần Thị B', '1992', '', '', '', '', '', '', 'Nữ',
            '012345678902', '0911000000', 'Hải Phòng', 'Hải Phòng', 'Phường Hồng Bàng', '',
            106.67, 20.85, 'Sốt xuất huyết', '', '', '', '', '', '', '', '', '', '', '', '', '',
            '10/07/2026', '', '', '', '', '', '', 'Trạm y tế', 'Hải Phòng', '', 'Đã duyệt',
            '11/07/2026 08:00', '', '', ''
        ])
        wb.save(case_file)
        corrupt_sheet_dimension(case_file)
        detected = core.detect_excel(case_file)
        assert detected[0] == 'case'
        summary = core.import_excel(case_file, db)
        assert summary.inserted == 1


# --- shift_iso_week / is_valid_iso_week ------------------------------------------------------
# Dùng cho dashboard chọn tuần khác để xem xã nào chưa nộp (mặc định tuần TRƯỚC, xem
# webapp/routers/dashboard.py).

def test_shift_iso_week_moves_back_one_week():
    assert core.shift_iso_week("2026-W31", -1) == "2026-W30"


def test_shift_iso_week_moves_forward_one_week():
    assert core.shift_iso_week("2026-W30", 1) == "2026-W31"


def test_shift_iso_week_crosses_year_boundary():
    # Tuần 1/2026 lùi 1 tuần phải ra đúng tuần cuối cùng của năm 2025 theo ISO-8601 (2025 chỉ có
    # 52 tuần ISO — 31/12/2025 đã thuộc tuần 1/2026).
    assert core.shift_iso_week("2026-W01", -1) == "2025-W52"


def test_shift_iso_week_returns_input_unchanged_when_invalid():
    assert core.shift_iso_week("khong-hop-le", -1) == "khong-hop-le"
    assert core.shift_iso_week("", -1) == ""


def test_is_valid_iso_week():
    assert core.is_valid_iso_week("2026-W31") is True
    assert core.is_valid_iso_week("2026-31") is False
    assert core.is_valid_iso_week("") is False
    assert core.is_valid_iso_week("khong-hop-le") is False
