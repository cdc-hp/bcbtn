"""Kiểm thử kiểm tra/đồng nhất định dạng ngày tháng (nhập + dữ liệu cũ) và xuất Excel dd/MM/yyyy."""

from __future__ import annotations

import sqlite3
import tempfile
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

import core


def make_excel(path: Path, fields, rows, sheet="Disease Cases"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append([label for label, _ in fields])
    for values in rows:
        ws.append([values.get(key, "") for _, key in fields])
    wb.save(path)


# ---------- Nhập file có định dạng ngày LẪN LỘN giữa các dòng/trường (kịch bản thực tế) --------

def test_import_unifies_mixed_date_formats_across_rows():
    """Mô phỏng dữ liệu thực tế do nhiều người/nhiều nguồn nhập tay — mỗi dòng dùng một kiểu
    định dạng ngày khác nhau cho CÙNG một trường. Toàn bộ phải quy về đúng một chuẩn lưu trữ
    (ISO) bất kể định dạng gốc là gì, để hiển thị/xuất Excel sau này luôn nhất quán."""
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        db = root / "test.db"
        core.BACKUP_DIR = root / "backups"
        file = root / "mixed_dates.xlsx"
        rows = [
            {
                "case_code": "MIX-01", "full_name": "Nguyễn Văn A", "main_diagnosis": "Sốt xuất huyết",
                "onset_date": "10/07/2026", "admission_date": "2026-07-11",
                "discharge_or_death_date": "15-07-2026", "sample_date": "12.07.2026",
                "report_datetime": "11/07/2026 08:30", "birth_date_raw": "15/05/1990",
            },
            {
                "case_code": "MIX-02", "full_name": "Trần Thị B", "main_diagnosis": "Sốt xuất huyết",
                "onset_date": "2026-07-10", "admission_date": "11-07-2026",
                "discharge_or_death_date": "2026/07/15", "sample_date": "12/07/2026",
                "report_datetime": "2026-07-11 08:30", "birth_date_raw": "1985",
            },
            {
                "case_code": "MIX-03", "full_name": "Lê Văn C", "main_diagnosis": "Sốt xuất huyết",
                "onset_date": "10.07.2026", "admission_date": "khong ro ngay",
                "report_datetime": "11.07.2026 08:30", "birth_date_raw": "05/1992",
            },
        ]
        make_excel(file, core.CASE_FIELDS, rows)
        summary = core.import_excel(file, db)
        assert summary.inserted == 3

        records, _ = core.query_records("case", db_path=db, page_size=10)
        by_code = {r["case_code"]: r for r in records}

        # Cùng một ngày thật, nhập bằng 5 cách viết khác nhau -> phải lưu ra CÙNG một chuỗi ISO.
        assert by_code["MIX-01"]["onset_date"] == "2026-07-10"
        assert by_code["MIX-02"]["onset_date"] == "2026-07-10"
        assert by_code["MIX-03"]["onset_date"] == "2026-07-10"
        assert by_code["MIX-01"]["admission_date"] == "2026-07-11"
        assert by_code["MIX-02"]["admission_date"] == "2026-07-11"
        assert by_code["MIX-01"]["discharge_or_death_date"] == "2026-07-15"
        assert by_code["MIX-02"]["discharge_or_death_date"] == "2026-07-15"
        assert by_code["MIX-01"]["sample_date"] == "2026-07-12"
        assert by_code["MIX-02"]["sample_date"] == "2026-07-12"
        assert by_code["MIX-01"]["report_datetime"] == "2026-07-11 08:30"
        assert by_code["MIX-02"]["report_datetime"] == "2026-07-11 08:30"
        assert by_code["MIX-03"]["report_datetime"] == "2026-07-11 08:30"
        assert by_code["MIX-01"]["birth_date_raw"] == "1990-05-15"

        # Không parse được ngày đầy đủ -> giữ nguyên văn, KHÔNG bị coi là lỗi (chỉ có năm/tháng-năm).
        assert by_code["MIX-02"]["birth_date_raw"] == "1985"
        assert by_code["MIX-03"]["birth_date_raw"] == "05/1992"

        # Giá trị KHÔNG phải ngày thật ("khong ro ngay") giữ nguyên văn và PHẢI bị cảnh báo.
        assert by_code["MIX-03"]["admission_date"] == "khong ro ngay"
        issues = core.list_quality_issues(entity_type="case", db_path=db)
        date_issues = [i for i in issues if i["issue_type"] == "Định dạng ngày không nhận diện được"]
        assert len(date_issues) == 1
        assert "Ngày nhập viện" in date_issues[0]["description"]
        assert "khong ro ngay" in date_issues[0]["description"]

        # Hiển thị Web: mọi giá trị ISO hợp lệ phải ra CÙNG dd/MM/yyyy bất kể định dạng gốc.
        for code in ("MIX-01", "MIX-02", "MIX-03"):
            display = core.format_record_dates(by_code[code])
            assert display["onset_date"] == "10/07/2026"
            assert display["report_datetime"] == "11/07/2026 08:30"
        assert core.format_record_dates(by_code["MIX-01"])["admission_date"] == "11/07/2026"
        assert core.format_record_dates(by_code["MIX-03"])["admission_date"] == "khong ro ngay"

        # Xuất Excel: mọi ngày hợp lệ phải ra cell date thật, cùng number_format dd/mm/yyyy.
        out = root / "out.xlsx"
        core.export_filtered_records(out, "case", db_path=db)
        wb2 = load_workbook(out)
        ws2 = wb2.active
        headers = [c.value for c in ws2[1]]
        onset_col = headers.index("Ngày khởi phát") + 1
        for r in range(2, 5):
            cell = ws2.cell(row=r, column=onset_col)
            assert cell.value.date() == date(2026, 7, 10)
            assert cell.number_format == "dd/mm/yyyy"


# ---------- parse_date_value: định dạng mở rộng ----------

def test_parse_date_value_accepts_dash_and_dot_separators():
    assert core.parse_date_value("10-07-2026") == "2026-07-10"
    assert core.parse_date_value("10.07.2026") == "2026-07-10"
    assert core.parse_date_value("2026/07/10") == "2026-07-10"
    assert core.parse_date_value("10-07-2026 08:30", with_time=True) == "2026-07-10 08:30"


def test_parse_date_value_falls_back_to_original_text_when_unrecognized():
    assert core.parse_date_value("khong phai ngay") == "khong phai ngay"


# ---------- Kiểm tra định dạng khi nhập (data_quality_issues) ----------

def test_import_flags_invalid_date_format_as_quality_issue():
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        db = root / "test.db"
        core.BACKUP_DIR = root / "backups"
        file = root / "cases.xlsx"
        rows = [{
            "full_name": "Nguyễn Văn A", "case_code": "CA-1", "main_diagnosis": "Sốt xuất huyết",
            "onset_date": "khong phai ngay", "report_datetime": "10/07/2026 08:00",
        }]
        make_excel(file, core.CASE_FIELDS, rows)
        core.import_excel(file, db)
        issues = core.list_quality_issues(entity_type="case", db_path=db)
        assert any(i["issue_type"] == "Định dạng ngày không nhận diện được" for i in issues)


def test_import_does_not_flag_valid_date_formats():
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        db = root / "test.db"
        core.BACKUP_DIR = root / "backups"
        file = root / "cases.xlsx"
        rows = [{
            "full_name": "Nguyễn Văn A", "case_code": "CA-1", "main_diagnosis": "Sốt xuất huyết",
            "onset_date": "10/07/2026", "report_datetime": "10/07/2026 08:00",
        }]
        make_excel(file, core.CASE_FIELDS, rows)
        core.import_excel(file, db)
        issues = core.list_quality_issues(entity_type="case", db_path=db)
        assert not any(i["issue_type"] == "Định dạng ngày không nhận diện được" for i in issues)


def test_import_does_not_flag_year_only_birth_date():
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        db = root / "test.db"
        core.BACKUP_DIR = root / "backups"
        file = root / "cases.xlsx"
        rows = [{
            "full_name": "Nguyễn Văn A", "case_code": "CA-1", "main_diagnosis": "Sốt xuất huyết",
            "onset_date": "10/07/2026", "report_datetime": "10/07/2026 08:00", "birth_date_raw": "1990",
        }]
        make_excel(file, core.CASE_FIELDS, rows)
        core.import_excel(file, db)
        issues = core.list_quality_issues(entity_type="case", db_path=db)
        assert not any(i["issue_type"] == "Định dạng ngày không nhận diện được" for i in issues)
        record = core.query_records("case", db_path=db)[0][0]
        assert record["birth_date_raw"] == "1990"
        assert record["birth_year"] == 1990


def test_import_normalizes_full_birth_date_to_iso():
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        db = root / "test.db"
        core.BACKUP_DIR = root / "backups"
        file = root / "cases.xlsx"
        rows = [{
            "full_name": "Nguyễn Văn A", "case_code": "CA-1", "main_diagnosis": "Sốt xuất huyết",
            "onset_date": "10/07/2026", "report_datetime": "10/07/2026 08:00", "birth_date_raw": "15-05-1990",
        }]
        make_excel(file, core.CASE_FIELDS, rows)
        core.import_excel(file, db)
        record = core.query_records("case", db_path=db)[0][0]
        assert record["birth_date_raw"] == "1990-05-15"
        assert record["birth_year"] == 1990


# ---------- Quét & chuẩn hóa dữ liệu cũ (normalize_stored_dates) ----------

def _insert_legacy_case(db: Path, onset_date_raw: str) -> int:
    """Chèn thẳng 1 dòng qua sqlite3, mô phỏng dữ liệu cũ có onset_date lưu KHÔNG chuẩn ISO —
    trạng thái mà normalize_stored_dates cần quét ra và sửa lại. Đóng kết nối tường minh (Windows
    giữ khóa file .db nếu chỉ dùng sqlite3.Connection làm context manager — xem core._connect)."""
    core.init_db(db)
    now = datetime.now().isoformat(sep=" ", timespec="seconds")
    conn = sqlite3.connect(db)
    try:
        cur = conn.execute(
            """INSERT INTO cases (full_name, case_code, main_diagnosis, onset_date, report_datetime,
               source_file, imported_at, row_hash)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            ("Nguyễn Văn A", "CA-LEGACY", "Sốt xuất huyết", onset_date_raw, "2026-07-10 08:00",
             "legacy.xlsx", now, f"legacy-hash-{onset_date_raw}"),
        )
        conn.commit()
        return int(cur.lastrowid)
    finally:
        conn.close()


def test_normalize_stored_dates_fixes_legacy_bad_format():
    with tempfile.TemporaryDirectory() as tmp:
        db = Path(tmp) / "test.db"
        record_id = _insert_legacy_case(db, "10-07-2026")
        result = core.normalize_stored_dates(db_path=db, actor="tester")
        assert result["updated_records"] == 1
        assert result["updated_fields"] == 1
        record = core.get_record("case", record_id, db_path=db)
        assert record["onset_date"] == "2026-07-10"


def test_normalize_stored_dates_is_idempotent():
    with tempfile.TemporaryDirectory() as tmp:
        db = Path(tmp) / "test.db"
        _insert_legacy_case(db, "10-07-2026")
        core.normalize_stored_dates(db_path=db, actor="tester")
        second = core.normalize_stored_dates(db_path=db, actor="tester")
        assert second["updated_records"] == 0


def test_normalize_stored_dates_skips_already_correct_records():
    with tempfile.TemporaryDirectory() as tmp:
        db = Path(tmp) / "test.db"
        _insert_legacy_case(db, "2026-07-10")  # đã đúng chuẩn ISO sẵn
        result = core.normalize_stored_dates(db_path=db, actor="tester")
        assert result["updated_records"] == 0


def test_normalize_stored_dates_recomputes_row_hash():
    with tempfile.TemporaryDirectory() as tmp:
        db = Path(tmp) / "test.db"
        record_id = _insert_legacy_case(db, "10-07-2026")
        conn = sqlite3.connect(db)
        before = conn.execute("SELECT row_hash FROM cases WHERE id=?", (record_id,)).fetchone()[0]
        conn.close()
        core.normalize_stored_dates(db_path=db, actor="tester")
        conn = sqlite3.connect(db)
        after = conn.execute("SELECT row_hash FROM cases WHERE id=?", (record_id,)).fetchone()[0]
        conn.close()
        assert before != after


def test_normalize_stored_dates_refreshes_quality_issues():
    """Bản ghi có onset_date lưu sai định dạng -> đang mang cảnh báo "Định dạng ngày không nhận
    diện được"; sau khi chuẩn hóa, cảnh báo đó phải biến mất."""
    with tempfile.TemporaryDirectory() as tmp:
        db = Path(tmp) / "test.db"
        record_id = _insert_legacy_case(db, "10-07-2026")
        core.init_db(db)
        conn = sqlite3.connect(db)
        conn.execute(
            """INSERT INTO data_quality_issues
               (entity_type, entity_id, source_file, source_row, severity, issue_type, description, created_at)
               VALUES ('case', ?, 'legacy.xlsx', 1, 'warning', 'Định dạng ngày không nhận diện được', 'cu', '2026-01-01')""",
            (record_id,),
        )
        conn.commit()
        conn.close()
        core.normalize_stored_dates(db_path=db, actor="tester")
        issues = core.list_quality_issues(entity_type="case", entity_id=record_id, db_path=db)
        assert not any(i["issue_type"] == "Định dạng ngày không nhận diện được" for i in issues)


# ---------- to_excel_date_value ----------

def test_to_excel_date_value_converts_date_field():
    assert core.to_excel_date_value("onset_date", "2026-07-10") == date(2026, 7, 10)


def test_to_excel_date_value_converts_datetime_field():
    assert core.to_excel_date_value("report_datetime", "2026-07-10 08:30") == datetime(2026, 7, 10, 8, 30)


def test_to_excel_date_value_converts_full_birth_date():
    assert core.to_excel_date_value("birth_date_raw", "1990-05-15") == date(1990, 5, 15)


def test_to_excel_date_value_keeps_year_only_birth_date():
    assert core.to_excel_date_value("birth_date_raw", "1990") == "1990"


def test_to_excel_date_value_keeps_non_date_column_unchanged():
    assert core.to_excel_date_value("full_name", "Nguyễn Văn A") == "Nguyễn Văn A"


# ---------- format_date_for_display / format_record_dates (giao diện Web) ----------

def test_format_date_for_display_date_field():
    assert core.format_date_for_display("onset_date", "2026-07-10") == "10/07/2026"


def test_format_date_for_display_datetime_field():
    assert core.format_date_for_display("report_datetime", "2026-07-10 08:30") == "10/07/2026 08:30"


def test_format_date_for_display_full_birth_date():
    assert core.format_date_for_display("birth_date_raw", "1990-05-15") == "15/05/1990"


def test_format_date_for_display_keeps_year_only_birth_date():
    assert core.format_date_for_display("birth_date_raw", "1990") == "1990"


def test_format_date_for_display_keeps_non_date_column_unchanged():
    assert core.format_date_for_display("full_name", "Nguyễn Văn A") == "Nguyễn Văn A"


def test_format_date_for_display_keeps_unparseable_date_unchanged():
    assert core.format_date_for_display("onset_date", "khong phai ngay") == "khong phai ngay"


def test_format_record_dates_only_touches_date_fields():
    record = {
        "id": 1, "full_name": "Nguyễn Văn A", "onset_date": "2026-07-10",
        "report_datetime": "2026-07-10 08:30", "birth_date_raw": "1990",
    }
    formatted = core.format_record_dates(record)
    assert formatted["onset_date"] == "10/07/2026"
    assert formatted["report_datetime"] == "10/07/2026 08:30"
    assert formatted["birth_date_raw"] == "1990"
    assert formatted["full_name"] == "Nguyễn Văn A"
    assert formatted["id"] == 1
    # Không đổi record gốc
    assert record["onset_date"] == "2026-07-10"


# ---------- format_timestamp_for_display (mốc thời gian hệ thống: nhập, nộp file) ------------

def test_format_timestamp_for_display_with_seconds():
    assert core.format_timestamp_for_display("2026-07-10 09:12:19") == "10/07/2026 09:12"


def test_format_timestamp_for_display_without_seconds():
    assert core.format_timestamp_for_display("2026-07-10 09:12") == "10/07/2026 09:12"


def test_format_timestamp_for_display_keeps_empty_unchanged():
    assert core.format_timestamp_for_display("") == ""
    assert core.format_timestamp_for_display(None) is None


def test_format_timestamp_for_display_keeps_unparseable_unchanged():
    assert core.format_timestamp_for_display("khong phai thoi gian") == "khong phai thoi gian"


def test_to_excel_date_value_keeps_empty_value():
    assert core.to_excel_date_value("onset_date", "") == ""
    assert core.to_excel_date_value("onset_date", None) is None


# ---------- Xuất Excel: dd/mm/yyyy ----------

def test_export_rows_xlsx_uses_dd_mm_yyyy_number_format():
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "out.xlsx"
        core.export_rows(
            path, ["Ngày khởi phát", "Thời gian báo cáo"],
            [[date(2026, 7, 10), datetime(2026, 7, 10, 8, 30)]],
        )
        wb = load_workbook(path)
        ws = wb.active
        onset_cell = ws.cell(row=2, column=1)
        report_cell = ws.cell(row=2, column=2)
        # openpyxl luôn đọc lại cell định dạng ngày dưới dạng datetime (kể cả khi ghi vào là
        # date thuần) — chỉ so phần ngày cho cột date-only.
        assert onset_cell.value.date() == date(2026, 7, 10)
        assert onset_cell.number_format == "dd/mm/yyyy"
        assert report_cell.value == datetime(2026, 7, 10, 8, 30)
        assert report_cell.number_format == "dd/mm/yyyy hh:mm"


def test_export_rows_csv_formats_date_as_dd_mm_yyyy_text():
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "out.csv"
        core.export_rows(path, ["Ngày khởi phát"], [[date(2026, 7, 10)]])
        text = path.read_text(encoding="utf-8-sig")
        assert "10/07/2026" in text
        assert "2026-07-10" not in text


def test_export_filtered_records_case_dates_are_dd_mm_yyyy():
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        db = root / "test.db"
        core.BACKUP_DIR = root / "backups"
        file = root / "cases.xlsx"
        rows = [{
            "full_name": "Nguyễn Văn A", "case_code": "CA-1", "main_diagnosis": "Sốt xuất huyết",
            "onset_date": "10/07/2026", "report_datetime": "10/07/2026 08:00",
        }]
        make_excel(file, core.CASE_FIELDS, rows)
        core.import_excel(file, db)

        out = root / "out.xlsx"
        core.export_filtered_records(out, "case", db_path=db)
        wb = load_workbook(out)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        onset_col = headers.index("Ngày khởi phát") + 1
        cell = ws.cell(row=2, column=onset_col)
        assert cell.value.date() == date(2026, 7, 10)
        assert cell.number_format == "dd/mm/yyyy"
