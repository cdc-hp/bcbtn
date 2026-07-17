from __future__ import annotations

import base64
import http.client
import io
import json
import tempfile
import threading
from datetime import datetime, timedelta
from http.server import BaseHTTPRequestHandler, HTTPServer
from pathlib import Path
from unittest.mock import patch
from urllib.error import HTTPError
from urllib.request import Request, urlopen

import lan_server

from openpyxl import Workbook, load_workbook

import core
import remote_core
import secondary_sync
from deployment_config import DeploymentConfig
from lan_server import LanServerController


def make_excel(path: Path, fields, rows, sheet="Disease Cases"):
    wb = Workbook(); ws = wb.active; ws.title = sheet
    ws.append([label for label, _ in fields])
    for values in rows:
        ws.append([values.get(key, "") for _, key in fields])
    wb.save(path)


def make_excel_bytes(fields, rows, sheet="Disease Cases") -> bytes:
    wb = Workbook(); ws = wb.active; ws.title = sheet
    ws.append([label for label, _ in fields])
    for values in rows:
        ws.append([values.get(key, "") for _, key in fields])
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


BASE_CASE = {
    "full_name": "Nguyễn Văn A", "birth_date_raw": "01/01/1990", "gender": "Nam",
    "phone": "0901234567", "main_diagnosis": "Sốt xuất huyết Dengue",
    "onset_date": "10/07/2026", "report_datetime": "11/07/2026 08:00", "reporting_unit": "Trạm Y tế",
}


def test_case_duplicate_default_criteria_matches_code_and_national_id():
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp); db = root / "test.db"; core.BACKUP_DIR = root / "backups"
        file = root / "cases.xlsx"
        rows = [
            dict(BASE_CASE, case_code="CA-100", national_id="123456789012", commune="Xã A"),
            dict(BASE_CASE, case_code="CA-100", national_id="123456789012", commune="Xã A", full_name="Nguyễn Văn A (2)"),
        ]
        make_excel(file, core.CASE_FIELDS, rows)
        assert core.import_excel(file, db).inserted == 2
        groups = core.find_duplicate_groups("case", db_path=db)
        assert len(groups) == 1
        assert groups[0]["confidence"] == "Trùng chắc chắn"
        assert set(groups[0]["case_codes"]) == {"CA-100"}


def test_case_duplicate_no_match_when_criteria_not_selected():
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp); db = root / "test.db"; core.BACKUP_DIR = root / "backups"
        file = root / "cases.xlsx"
        rows = [dict(BASE_CASE, case_code="CA-1"), dict(BASE_CASE, case_code="CA-2")]
        make_excel(file, core.CASE_FIELDS, rows)
        assert core.import_excel(file, db).inserted == 2
        groups = core.find_duplicate_groups("case", db_path=db, criteria={"enabled": ["national_id"]})
        assert groups == []


def test_case_duplicate_name_similar_detects_pairs_within_same_commune():
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp); db = root / "test.db"; core.BACKUP_DIR = root / "backups"
        file = root / "cases.xlsx"
        rows = [
            dict(BASE_CASE, case_code="CA-N1", full_name="Nguyễn Văn An", phone="0911111111",
                 birth_date_raw="01/01/1980", commune="Xã A"),
            dict(BASE_CASE, case_code="CA-N2", full_name="Nguyễn Văn Anh", phone="0922222222",
                 birth_date_raw="02/02/1990", commune="Xã A"),
        ]
        make_excel(file, core.CASE_FIELDS, rows)
        assert core.import_excel(file, db).inserted == 2
        groups = core.find_duplicate_groups("case", db_path=db, criteria={"enabled": ["name_similar"]})
        assert len(groups) == 1
        assert any("gần giống" in c for c in groups[0]["matched_criteria"])


def test_case_duplicate_onset_near_detects_pairs_within_same_commune_only():
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp); db = root / "test.db"; core.BACKUP_DIR = root / "backups"
        file = root / "cases.xlsx"
        rows = [
            dict(BASE_CASE, case_code="CA-O1", full_name="Lê Thị Hoa", phone="0933333333",
                 onset_date="10/07/2026", commune="Xã A"),
            dict(BASE_CASE, case_code="CA-O2", full_name="Phạm Văn Bình", phone="0944444444",
                 onset_date="12/07/2026", commune="Xã A"),
            dict(BASE_CASE, case_code="CA-O3", full_name="Đỗ Thị Mai", phone="0955555555",
                 onset_date="11/07/2026", commune="Xã B"),
        ]
        make_excel(file, core.CASE_FIELDS, rows)
        assert core.import_excel(file, db).inserted == 3
        groups = core.find_duplicate_groups(
            "case", db_path=db, criteria={"enabled": ["onset_near"], "onset_max_days": 3}
        )
        # Xã A có 2 ca lệch ngày khởi phát trong ngưỡng -> phát hiện được nhờ bucket theo xã.
        assert len(groups) == 1
        assert set(groups[0]["case_codes"]) == {"CA-O1", "CA-O2"}
        # Ca ở Xã B không được so với Xã A dù cũng lệch ngày trong ngưỡng — giới hạn đã biết
        # (name_similar/onset_near chỉ so trong cùng xã).
        assert "CA-O3" not in groups[0]["case_codes"]


def test_export_cases_by_commune_resolves_cross_commune_by_latest_admission():
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp); db = root / "test.db"; core.BACKUP_DIR = root / "backups"
        file = root / "cases.xlsx"
        rows = [
            dict(BASE_CASE, case_code="CA-1", national_id="111111111111", commune="Xã A", admission_date="10/07/2026"),
            dict(BASE_CASE, case_code="CA-1-B", national_id="111111111111", commune="Xã B", admission_date="15/07/2026"),
            dict(BASE_CASE, case_code="CA-2", full_name="Trần Thị B", commune="Xã C"),
        ]
        make_excel(file, core.CASE_FIELDS, rows)
        assert core.import_excel(file, db).inserted == 3
        out = root / "theo_xa.xlsx"
        result = core.export_cases_by_commune(out, db_path=db)
        assert result["case_count"] == 3
        assert result["cross_commune_group_count"] == 1
        wb = load_workbook(out)
        assert "Tong_hop" in wb.sheetnames
        assert "Xã B" in wb.sheetnames
        assert "Xã A" not in wb.sheetnames
        assert "Xã C" in wb.sheetnames
        commune_b_codes = {row[2].value for row in wb["Xã B"].iter_rows(min_row=2)}
        assert commune_b_codes == {"CA-1", "CA-1-B"}


def test_export_cases_by_commune_dedup_scope_covers_all_records():
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp); db = root / "test.db"; core.BACKUP_DIR = root / "backups"
        file = root / "cases.xlsx"
        rows = [dict(BASE_CASE, case_code=f"CA-{i}", commune="Xã A") for i in range(5)]
        make_excel(file, core.CASE_FIELDS, rows)
        assert core.import_excel(file, db).inserted == 5
        out = root / "theo_xa.xlsx"
        with patch("core._find_case_duplicate_groups", wraps=core._find_case_duplicate_groups) as spy:
            core.export_cases_by_commune(out, db_path=db)
            max_records_used = spy.call_args.args[2]
            assert max_records_used >= 5, (
                "phạm vi dò trùng khi xuất theo xã phải phủ hết số ca thực tế, "
                "không được cố định thấp hơn tổng số ca"
            )


def test_import_queue_submit_list_and_import():
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp); db = root / "test.db"
        core.BACKUP_DIR = root / "backups"; core.QUEUE_DIR = root / "queue"
        data = make_excel_bytes(core.CASE_FIELDS, [dict(BASE_CASE, case_code="CA-Q1")])
        submitted = core.queue_submit("Xã Gia Viên", "2026-W29", "danh_sach.xlsx", data, db_path=db)
        assert submitted["commune"] == "Xã Gia Viên"
        items = core.list_import_queue(db_path=db)
        assert len(items) == 1 and items[0]["status"] == "cho_nhap"
        imported = core.import_queue_item(items[0]["id"], db_path=db)
        assert imported["inserted"] == 1
        items_after = core.list_import_queue(db_path=db)
        assert items_after[0]["status"] == "da_nhap"
        try:
            core.import_queue_item(items[0]["id"], db_path=db)
            assert False, "phải báo lỗi khi nhập lại mục đã nhập"
        except ValueError:
            pass


def test_queue_submit_same_second_does_not_overwrite_file():
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp); db = root / "test.db"
        core.BACKUP_DIR = root / "backups"; core.QUEUE_DIR = root / "queue"
        data_a = make_excel_bytes(core.CASE_FIELDS, [dict(BASE_CASE, case_code="CA-DUP-A")])
        data_b = make_excel_bytes(core.CASE_FIELDS, [dict(BASE_CASE, case_code="CA-DUP-B")])
        first = core.queue_submit("Xã Gia Viên", "2026-W29", "danh_sach.xlsx", data_a, db_path=db)
        second = core.queue_submit("Xã Gia Viên", "2026-W29", "danh_sach.xlsx", data_b, db_path=db)
        items = {item["id"]: item for item in core.list_import_queue(db_path=db)}
        path_a = Path(items[first["queue_id"]]["file_path"])
        path_b = Path(items[second["queue_id"]]["file_path"])
        assert path_a != path_b
        assert path_a.exists() and path_b.exists()
        assert path_a.read_bytes() == data_a
        assert path_b.read_bytes() == data_b


def test_import_queue_item_concurrent_calls_only_import_once():
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp); db = root / "test.db"
        core.BACKUP_DIR = root / "backups"; core.QUEUE_DIR = root / "queue"
        data = make_excel_bytes(core.CASE_FIELDS, [dict(BASE_CASE, case_code="CA-RACE")])
        submitted = core.queue_submit("Xã Gia Viên", "2026-W29", "danh_sach.xlsx", data, db_path=db)
        queue_id = submitted["queue_id"]

        results: list[tuple[bool, str]] = []
        results_lock = threading.Lock()
        start_barrier = threading.Barrier(2)

        def worker():
            start_barrier.wait()
            try:
                core.import_queue_item(queue_id, db_path=db)
                with results_lock:
                    results.append((True, ""))
            except ValueError as exc:
                with results_lock:
                    results.append((False, str(exc)))

        threads = [threading.Thread(target=worker) for _ in range(2)]
        for t in threads:
            t.start()
        for t in threads:
            t.join(timeout=10)

        successes = [r for r in results if r[0]]
        failures = [r for r in results if not r[0]]
        assert len(successes) == 1, f"chỉ một lần gọi được thành công, nhận: {results}"
        assert len(failures) == 1
        assert core.dashboard_stats(db)["case_records"] == 1
        items = core.list_import_queue(db_path=db)
        assert items[0]["status"] == "da_nhap"


def test_lan_server_queue_endpoints():
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        core.DATA_DIR = root / "data"; core.DATA_DIR.mkdir()
        core.DB_PATH = core.DATA_DIR / "test.db"
        core.BACKUP_DIR = root / "backups"; core.QUEUE_DIR = root / "queue"
        cfg = DeploymentConfig(mode="server", server_host="127.0.0.1", server_port=0, password="")
        ctrl = LanServerController(cfg)
        ctrl.start()
        addr = f"http://127.0.0.1:{ctrl.port}"
        try:
            html = urlopen(addr + "/xa", timeout=5).read().decode("utf-8")
            assert "Nộp danh sách ca bệnh" in html
            data = make_excel_bytes(core.CASE_FIELDS, [dict(BASE_CASE, case_code="CA-LAN")])
            body = json.dumps({
                "commune": "Xã Gia Viên", "week": "2026-W29", "file_name": "ds.xlsx",
                "content_base64": base64.b64encode(data).decode("ascii"),
            }).encode()
            req = Request(addr + "/queue/submit", data=body, headers={"Content-Type": "application/json"}, method="POST")
            resp = json.loads(urlopen(req, timeout=10).read().decode("utf-8"))
            assert resp["ok"]
            queue_id = resp["result"]["queue_id"]

            req2 = Request(addr + "/rpc", data=json.dumps({"function": "list_import_queue"}).encode(), headers={"Content-Type": "application/json"}, method="POST")
            resp2 = json.loads(urlopen(req2, timeout=10).read().decode("utf-8"))
            assert resp2["ok"] and len(resp2["result"]) == 1

            req3 = Request(addr + "/rpc", data=json.dumps({"function": "import_queue_item", "args": [queue_id]}).encode(), headers={"Content-Type": "application/json"}, method="POST")
            resp3 = json.loads(urlopen(req3, timeout=10).read().decode("utf-8"))
            assert resp3["ok"] and resp3["result"]["inserted"] == 1
        finally:
            ctrl.stop()


class _FakeSecondaryServerHandler(BaseHTTPRequestHandler):
    pending = [{
        "row": 2, "commune": "Xã Đông Hải", "week": "2026-W29", "file_name": "xa_dong_hai.xlsx",
        "submitted_by": "Y tá B",
        "content_base64": base64.b64encode(
            make_excel_bytes(core.CASE_FIELDS, [dict(BASE_CASE, case_code="CA-SEC")])
        ).decode("ascii"),
    }]
    synced: list[int] = []

    def log_message(self, *args):
        return

    def do_POST(self):
        length = int(self.headers.get("Content-Length", "0"))
        payload = json.loads(self.rfile.read(length).decode())
        assert payload.get("key") == "s3cr3t", "khóa phải luôn gửi qua POST body, không qua query string"
        if payload.get("action") == "list_pending":
            body = json.dumps({"ok": True, "result": self.pending}).encode()
        elif payload.get("action") == "mark_synced":
            _FakeSecondaryServerHandler.synced.extend(payload.get("rows", []))
            body = json.dumps({"ok": True, "result": {"marked": payload.get("rows", [])}}).encode()
        else:
            body = json.dumps({"ok": False, "error": "unknown action"}).encode()
        self.send_response(200); self.send_header("Content-Type", "application/json"); self.end_headers(); self.wfile.write(body)


def test_secondary_sync_pulls_pending_into_local_queue():
    server = HTTPServer(("127.0.0.1", 0), _FakeSecondaryServerHandler)
    thread = threading.Thread(target=server.serve_forever, daemon=True); thread.start()
    try:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp); db = root / "test.db"
            core.BACKUP_DIR = root / "backups"; core.QUEUE_DIR = root / "queue"
            url = f"http://127.0.0.1:{server.server_port}/"
            result = secondary_sync.pull_secondary_queue(url, "s3cr3t", db_path=db)
            assert result == {"pending_count": 1, "pulled_count": 1, "errors": []}
            assert _FakeSecondaryServerHandler.synced == [2]
            items = core.list_import_queue(db_path=db)
            assert len(items) == 1
            assert items[0]["source"] == "server_phu"
            assert items[0]["commune"] == "Xã Đông Hải"
    finally:
        server.shutdown()


def test_remote_core_queue_proxies_against_live_server(monkeypatch):
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        core.DATA_DIR = root / "data"; core.DATA_DIR.mkdir()
        core.DB_PATH = core.DATA_DIR / "test.db"
        core.BACKUP_DIR = root / "backups"; core.QUEUE_DIR = root / "queue"
        cfg = DeploymentConfig(mode="server", server_host="127.0.0.1", server_port=0, password="")
        ctrl = LanServerController(cfg)
        ctrl.start()
        try:
            workstation_cfg = DeploymentConfig(mode="workstation", server_url=f"http://127.0.0.1:{ctrl.port}")
            monkeypatch.setattr(remote_core, "load_config", lambda: workstation_cfg)
            data = make_excel_bytes(core.CASE_FIELDS, [dict(BASE_CASE, case_code="CA-REMOTE")])
            submitted = remote_core.queue_submit("Xã Gia Viên", "2026-W29", "ds.xlsx", data, submitted_by="Y tá C")
            assert submitted["commune"] == "Xã Gia Viên"
            items = remote_core.list_import_queue()
            assert len(items) == 1 and items[0]["status"] == "cho_nhap"
            imported = remote_core.import_queue_item(items[0]["id"])
            assert imported["inserted"] == 1
            items_after = remote_core.list_import_queue(status="da_nhap")
            assert len(items_after) == 1
        finally:
            ctrl.stop()


def test_queue_submit_rate_limit_returns_429(monkeypatch):
    monkeypatch.setattr(lan_server, "QUEUE_SUBMIT_RATE_LIMIT", 2)
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        core.DATA_DIR = root / "data"; core.DATA_DIR.mkdir()
        core.DB_PATH = core.DATA_DIR / "test.db"
        core.BACKUP_DIR = root / "backups"; core.QUEUE_DIR = root / "queue"
        cfg = DeploymentConfig(mode="server", server_host="127.0.0.1", server_port=0, password="")
        ctrl = LanServerController(cfg)
        ctrl.start()
        addr = f"http://127.0.0.1:{ctrl.port}"
        try:
            data = make_excel_bytes(core.CASE_FIELDS, [dict(BASE_CASE, case_code="CA-RATE")])
            body = json.dumps({
                "commune": "Xã Gia Viên", "week": "2026-W29", "file_name": "ds.xlsx",
                "content_base64": base64.b64encode(data).decode("ascii"),
            }).encode()

            def submit():
                req = Request(addr + "/queue/submit", data=body, headers={"Content-Type": "application/json"}, method="POST")
                return urlopen(req, timeout=10)

            submit().read()
            submit().read()
            try:
                submit()
                assert False, "lần nộp thứ 3 trong cửa sổ giới hạn phải bị chặn (429)"
            except HTTPError as exc:
                assert exc.code == 429
        finally:
            ctrl.stop()


def test_auth_rejection_drains_body_before_next_request_on_keepalive():
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        core.DATA_DIR = root / "data"; core.DATA_DIR.mkdir()
        core.DB_PATH = core.DATA_DIR / "test.db"
        core.BACKUP_DIR = root / "backups"; core.QUEUE_DIR = root / "queue"
        cfg = DeploymentConfig(mode="server", server_host="127.0.0.1", server_port=0, password="secret")
        ctrl = LanServerController(cfg)
        ctrl.start()
        try:
            conn = http.client.HTTPConnection("127.0.0.1", ctrl.port, timeout=10)
            body = json.dumps({"function": "dashboard_stats", "args": [], "kwargs": {}}).encode()
            conn.request("POST", "/rpc", body=body, headers={
                "Content-Type": "application/json", "X-GSBTN-Password": "wrong",
            })
            resp1 = conn.getresponse()
            assert resp1.status == 401
            resp1.read()
            # Cùng một kết nối keep-alive: nếu request trên chưa đọc hết thân, request này sẽ bị
            # đọc lệch (server hiểu nhầm phần thân cũ là dòng đầu request mới).
            body2 = json.dumps({"function": "dashboard_stats", "args": [], "kwargs": {}}).encode()
            conn.request("POST", "/rpc", body=body2, headers={
                "Content-Type": "application/json", "X-GSBTN-Password": "secret",
            })
            resp2 = conn.getresponse()
            payload = json.loads(resp2.read().decode("utf-8"))
            assert resp2.status == 200
            assert payload["ok"] is True
            conn.close()
        finally:
            ctrl.stop()


def test_list_import_queue_pagination():
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp); db = root / "test.db"
        core.BACKUP_DIR = root / "backups"; core.QUEUE_DIR = root / "queue"
        for i in range(3):
            data = make_excel_bytes(core.CASE_FIELDS, [dict(BASE_CASE, case_code=f"CA-P{i}")])
            core.queue_submit("Xã Gia Viên", "2026-W29", f"ds{i}.xlsx", data, db_path=db)
        page1 = core.list_import_queue(db_path=db, limit=2, offset=0)
        page2 = core.list_import_queue(db_path=db, limit=2, offset=2)
        all_items = core.list_import_queue(db_path=db, limit=10)
        assert len(page1) == 2
        assert len(page2) == 1
        assert {item["id"] for item in page1} | {item["id"] for item in page2} == {item["id"] for item in all_items}


def test_archive_old_queue_files_removes_only_old_imported_files():
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp); db = root / "test.db"
        core.BACKUP_DIR = root / "backups"; core.QUEUE_DIR = root / "queue"
        data_old = make_excel_bytes(core.CASE_FIELDS, [dict(BASE_CASE, case_code="CA-OLD")])
        data_new = make_excel_bytes(core.CASE_FIELDS, [dict(BASE_CASE, case_code="CA-NEW")])
        old_submit = core.queue_submit("Xã A", "2026-W20", "cu.xlsx", data_old, db_path=db)
        new_submit = core.queue_submit("Xã A", "2026-W29", "moi.xlsx", data_new, db_path=db)
        core.import_queue_item(old_submit["queue_id"], db_path=db)
        core.import_queue_item(new_submit["queue_id"], db_path=db)
        old_date = (datetime.now() - timedelta(days=200)).isoformat(sep=" ", timespec="seconds")
        with core._connect(db) as conn:
            conn.execute("UPDATE import_queue SET imported_at=? WHERE id=?", (old_date, old_submit["queue_id"]))

        items_before = {item["id"]: item for item in core.list_import_queue(db_path=db)}
        old_path = Path(items_before[old_submit["queue_id"]]["file_path"])
        new_path = Path(items_before[new_submit["queue_id"]]["file_path"])
        assert old_path.exists() and new_path.exists()

        result = core.archive_old_queue_files(older_than_days=90, db_path=db)
        assert result["archived_count"] == 1
        assert not old_path.exists()
        assert new_path.exists()

        items_after = {item["id"]: item for item in core.list_import_queue(db_path=db)}
        assert items_after[old_submit["queue_id"]]["archived_at"] is not None
        assert items_after[new_submit["queue_id"]]["archived_at"] is None

        # Chạy lại lần 2 không đụng tới các mục đã dọn (idempotent, không lỗi).
        result2 = core.archive_old_queue_files(older_than_days=90, db_path=db)
        assert result2["archived_count"] == 0
