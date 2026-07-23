"""Kiểm thử Giai đoạn 5: /cdc/xuat-du-lieu (xuất theo bộ lọc, xuất ca bệnh theo xã) — xem
TASKS.md."""

from __future__ import annotations

from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

import core
import deployment_config
import duplicate_config


@pytest.fixture()
def client(tmp_path: Path, monkeypatch):
    monkeypatch.setattr(core, "DATA_DIR", tmp_path / "data")
    core.DATA_DIR.mkdir(parents=True, exist_ok=True)
    monkeypatch.setattr(core, "DB_PATH", core.DATA_DIR / "test.db")
    monkeypatch.setattr(core, "BACKUP_DIR", tmp_path / "backups")
    monkeypatch.setattr(core, "QUEUE_DIR", tmp_path / "queue")
    monkeypatch.setattr(deployment_config, "CONFIG_PATH", tmp_path / "deployment.json")
    monkeypatch.setattr(duplicate_config, "CONFIG_PATH", tmp_path / "duplicate_rules.json")
    monkeypatch.setattr(duplicate_config, "CRITERIA_CONFIG_PATH", tmp_path / "case_duplicate_criteria.json")

    import webapp.main as webapp_main
    return TestClient(webapp_main.app)


def _fresh_csrf(client: TestClient, path: str) -> str:
    client.get(path)
    return client.cookies.get("csrf_token", "")


def _login(client: TestClient, role: str = "admin") -> None:
    client.get("/cdc/setup")
    csrf = client.cookies.get("csrf_token", "")
    client.post("/cdc/setup", data={
        "username": "sa_admin", "display_name": "Super", "password": "matkhau123",
        "password_confirm": "matkhau123", "csrf_token": csrf,
    })
    if role != "super_admin":
        core.create_cdc_account("cdc_user", "matkhau123", role=role, must_change_password=False, db_path=core.DB_PATH)
        csrf2 = _fresh_csrf(client, "/cdc/login")
        client.post("/cdc/login", data={"username": "cdc_user", "password": "matkhau123", "csrf_token": csrf2})
    else:
        csrf2 = _fresh_csrf(client, "/cdc/login")
        client.post("/cdc/login", data={"username": "sa_admin", "password": "matkhau123", "csrf_token": csrf2})


def _seed_cases(tmp_path: Path, rows: list[dict]) -> None:
    wb = Workbook(); ws = wb.active; ws.title = "Disease Cases"
    ws.append([label for label, _ in core.CASE_FIELDS])
    for row in rows:
        full = {key: "" for _, key in core.CASE_FIELDS}
        full.update(row)
        ws.append([full.get(key, "") for _, key in core.CASE_FIELDS])
    path = tmp_path / "seed_export_cases.xlsx"
    wb.save(path)
    core.import_excel(path, core.DB_PATH)


def test_export_hub_requires_login(client: TestClient):
    resp = client.get("/cdc/xuat-du-lieu", follow_redirects=False)
    assert resp.status_code == 303


def test_export_hub_hides_actions_for_viewer(client: TestClient):
    _login(client, role=core.CDC_ROLE_VIEWER)
    resp = client.get("/cdc/xuat-du-lieu")
    assert resp.status_code == 200
    assert "Tải xuống" not in resp.text
    assert "Xuất theo xã</a>" not in resp.text


def test_export_filtered_downloads_xlsx(client: TestClient, tmp_path: Path):
    _login(client)
    _seed_cases(tmp_path, [
        {"case_code": "CA-1", "full_name": "Nguyễn Văn A", "commune": "Xã A"},
        {"case_code": "CA-2", "full_name": "Trần Thị B", "commune": "Xã B"},
    ])
    resp = client.get("/cdc/xuat-du-lieu/tai-ve", params={"entity": "case", "search": "CA-1", "fmt": "xlsx"})
    assert resp.status_code == 200
    assert resp.headers["content-disposition"].startswith("attachment")
    out_path = tmp_path / "downloaded.xlsx"
    out_path.write_bytes(resp.content)
    wb = load_workbook(out_path)
    ws = wb.active
    values = [cell.value for row in ws.iter_rows(min_row=2) for cell in row]
    assert "CA-1" in values and "CA-2" not in values


def test_export_filtered_requires_export_role(client: TestClient, tmp_path: Path):
    _login(client, role=core.CDC_ROLE_VIEWER)
    _seed_cases(tmp_path, [{"case_code": "CA-1", "full_name": "Nguyễn Văn A", "commune": "Xã A"}])
    resp = client.get("/cdc/xuat-du-lieu/tai-ve", params={"entity": "case"})
    assert resp.status_code == 403


def test_export_filtered_empty_redirects_with_error(client: TestClient):
    _login(client)
    resp = client.get("/cdc/xuat-du-lieu/tai-ve", params={"entity": "case"}, follow_redirects=False)
    assert resp.status_code == 303
    assert "err=" in resp.headers["location"]


def test_export_by_commune_downloads_workbook(client: TestClient, tmp_path: Path):
    _login(client)
    _seed_cases(tmp_path, [
        {"case_code": "CA-1", "full_name": "Nguyễn Văn A", "commune": "Xã A"},
        {"case_code": "CA-2", "full_name": "Trần Thị B", "commune": "Xã B"},
    ])
    resp = client.get("/cdc/xuat-du-lieu/theo-xa")
    assert resp.status_code == 200
    out_path = tmp_path / "theo_xa.xlsx"
    out_path.write_bytes(resp.content)
    wb = load_workbook(out_path)
    assert "Tong_hop" in wb.sheetnames
    assert "Xã A" in wb.sheetnames and "Xã B" in wb.sheetnames

    actions = core.list_audit_log(db_path=core.DB_PATH)
    assert any(a["action"] == "export_cases_by_commune" for a in actions)


def test_export_by_commune_requires_export_role(client: TestClient, tmp_path: Path):
    _login(client, role=core.CDC_ROLE_VIEWER)
    _seed_cases(tmp_path, [{"case_code": "CA-1", "full_name": "Nguyễn Văn A", "commune": "Xã A"}])
    resp = client.get("/cdc/xuat-du-lieu/theo-xa")
    assert resp.status_code == 403
