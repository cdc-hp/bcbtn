"""Kiểm thử Giai đoạn 3: POST /queue/submit (tương thích Code.gs) + /cdc/hang-doi (xem TASKS.md)."""

from __future__ import annotations

import base64
import io
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

import core
import deployment_config
from deployment_config import save_config


def make_excel_b64(case_code: str) -> str:
    wb = Workbook(); ws = wb.active; ws.title = "Disease Cases"
    ws.append([label for label, _ in core.CASE_FIELDS])
    row = {key: "" for _, key in core.CASE_FIELDS}
    row["case_code"] = case_code
    row["full_name"] = "Nguyễn Văn Test"
    row["commune"] = "Xã Gia Viên"
    ws.append([row.get(key, "") for _, key in core.CASE_FIELDS])
    buf = io.BytesIO(); wb.save(buf)
    return base64.b64encode(buf.getvalue()).decode("ascii")


@pytest.fixture()
def client(tmp_path: Path, monkeypatch):
    monkeypatch.setattr(core, "DATA_DIR", tmp_path / "data")
    core.DATA_DIR.mkdir(parents=True, exist_ok=True)
    monkeypatch.setattr(core, "DB_PATH", core.DATA_DIR / "test.db")
    monkeypatch.setattr(core, "BACKUP_DIR", tmp_path / "backups")
    monkeypatch.setattr(core, "QUEUE_DIR", tmp_path / "queue")
    monkeypatch.setattr(deployment_config, "CONFIG_PATH", tmp_path / "deployment.json")

    from webapp.services.rate_limit import queue_submit_limiter
    queue_submit_limiter._hits.clear()

    import webapp.main as webapp_main
    return TestClient(webapp_main.app)


def _set_gas_api_key(key: str) -> None:
    cfg = deployment_config.load_config()
    cfg.gas_api_key = key
    save_config(cfg)


def _login_as(client: TestClient, role: str, username: str = "cdc_user") -> None:
    client.get("/cdc/setup")
    csrf = client.cookies.get("csrf_token", "")
    client.post("/cdc/setup", data={
        "username": "sa_admin", "display_name": "Super", "password": "matkhau123",
        "password_confirm": "matkhau123", "csrf_token": csrf,
    })
    core.create_cdc_account(username, "matkhau123", role=role, must_change_password=False, db_path=core.DB_PATH)
    csrf2 = _fresh_csrf(client, "/cdc/login")
    client.post("/cdc/login", data={"username": username, "password": "matkhau123", "csrf_token": csrf2})


def _fresh_csrf(client: TestClient, path: str) -> str:
    client.get(path)
    return client.cookies.get("csrf_token", "")


# ---------- POST /queue/submit ----------

def test_submit_rejected_when_api_key_not_configured(client: TestClient):
    resp = client.post("/queue/submit", json={"commune": "Xã A", "week": "2026-W29", "content_base64": make_excel_b64("C1")})
    assert resp.status_code == 503
    assert resp.json()["ok"] is False


def test_submit_rejected_with_wrong_key(client: TestClient):
    _set_gas_api_key("khoa-dung")
    resp = client.post(
        "/queue/submit", json={"commune": "Xã A", "week": "2026-W29", "content_base64": make_excel_b64("C1")},
        headers={"X-GSBTN-Password": "sai"},
    )
    assert resp.status_code == 401


def test_submit_success_creates_queue_entry(client: TestClient):
    _set_gas_api_key("khoa-dung")
    resp = client.post(
        "/queue/submit",
        json={"commune": "Xã Gia Viên", "week": "2026-W29", "file_name": "ds.xlsx", "content_base64": make_excel_b64("C1"), "submitted_by": "GAS"},
        headers={"X-GSBTN-Password": "khoa-dung"},
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["ok"] is True
    assert body["result"]["duplicate"] is False
    assert len(core.list_import_queue(db_path=core.DB_PATH)) == 1


def test_submit_duplicate_returns_same_queue_id(client: TestClient):
    _set_gas_api_key("khoa-dung")
    payload = {"commune": "Xã Gia Viên", "week": "2026-W29", "file_name": "ds.xlsx", "content_base64": make_excel_b64("C1")}
    first = client.post("/queue/submit", json=payload, headers={"X-GSBTN-Password": "khoa-dung"}).json()
    second = client.post("/queue/submit", json=payload, headers={"X-GSBTN-Password": "khoa-dung"}).json()
    assert second["result"]["duplicate"] is True
    assert second["result"]["queue_id"] == first["result"]["queue_id"]


def test_submit_missing_commune_returns_400(client: TestClient):
    _set_gas_api_key("khoa-dung")
    resp = client.post(
        "/queue/submit", json={"commune": "", "week": "2026-W29", "content_base64": make_excel_b64("C1")},
        headers={"X-GSBTN-Password": "khoa-dung"},
    )
    assert resp.status_code == 400


def test_submit_rate_limited_after_threshold(client: TestClient):
    _set_gas_api_key("khoa-dung")
    for i in range(10):
        resp = client.post(
            "/queue/submit",
            json={"commune": "Xã Rate", "week": "2026-W29", "file_name": f"{i}.xlsx", "content_base64": make_excel_b64(f"C{i}")},
            headers={"X-GSBTN-Password": "khoa-dung"},
        )
        assert resp.status_code == 200
    over_limit = client.post(
        "/queue/submit",
        json={"commune": "Xã Rate", "week": "2026-W29", "file_name": "over.xlsx", "content_base64": make_excel_b64("C-over")},
        headers={"X-GSBTN-Password": "khoa-dung"},
    )
    assert over_limit.status_code == 429


# ---------- /cdc/hang-doi ----------

def test_queue_page_requires_login(client: TestClient):
    resp = client.get("/cdc/hang-doi", follow_redirects=False)
    assert resp.status_code == 303


def test_queue_page_lists_and_filters(client: TestClient):
    _login_as(client, core.CDC_ROLE_ADMIN)
    core.queue_submit("Xã A", "2026-W20", "a.xlsx", base64.b64decode(make_excel_b64("A1")), db_path=core.DB_PATH)
    core.queue_submit("Xã B", "2026-W21", "b.xlsx", base64.b64decode(make_excel_b64("B1")), db_path=core.DB_PATH)

    page = client.get("/cdc/hang-doi")
    assert page.status_code == 200
    assert "a.xlsx" in page.text and "b.xlsx" in page.text

    filtered = client.get("/cdc/hang-doi", params={"commune": "Xã A"})
    assert "a.xlsx" in filtered.text and "b.xlsx" not in filtered.text


def test_queue_page_shows_received_at_as_dd_mm_yyyy_hh_mm(client: TestClient):
    _login_as(client, core.CDC_ROLE_ADMIN)
    core.queue_submit("Xã A", "2026-W20", "a.xlsx", base64.b64decode(make_excel_b64("A1")), db_path=core.DB_PATH)
    item = core.list_import_queue(db_path=core.DB_PATH)[0]
    raw_received_at = item["received_at"]

    page = client.get("/cdc/hang-doi")
    assert raw_received_at not in page.text
    assert core.format_timestamp_for_display(raw_received_at) in page.text


# ---------- Xuất Excel ----------

def test_queue_page_has_export_button(client: TestClient):
    _login_as(client, core.CDC_ROLE_ADMIN)
    resp = client.get("/cdc/hang-doi")
    assert "/cdc/hang-doi/xuat" in resp.text
    assert "Xuất Excel" in resp.text


def test_export_queue_returns_xlsx_with_expected_rows(client: TestClient):
    _login_as(client, core.CDC_ROLE_ADMIN)
    core.queue_submit("Xã A", "2026-W20", "a.xlsx", base64.b64decode(make_excel_b64("A1")), db_path=core.DB_PATH)
    resp = client.get("/cdc/hang-doi/xuat")
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("application/")

    from openpyxl import load_workbook
    import tempfile
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "out.xlsx"
        out.write_bytes(resp.content)
        wb = load_workbook(out)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        assert headers == ["Xã", "Tuần", "File", "Nguồn", "Trạng thái", "Người nộp", "Nhận lúc", "Cảnh báo"]
        data_row = [c.value for c in ws[2]]
        assert data_row[0] == "Xã A"
        assert data_row[1] == "2026-W20"
        assert data_row[2] == "a.xlsx"
        assert data_row[3] == "Trực tiếp"
        assert data_row[7] in (None, "")
        assert data_row[4] == "Chờ nhập"
        assert "-" not in data_row[6].split(" ")[0]  # dd/MM/yyyy, không phải ISO


def test_export_queue_respects_commune_filter(client: TestClient):
    _login_as(client, core.CDC_ROLE_ADMIN)
    core.queue_submit("Xã A", "2026-W20", "a.xlsx", base64.b64decode(make_excel_b64("A1")), db_path=core.DB_PATH)
    core.queue_submit("Xã B", "2026-W20", "b.xlsx", base64.b64decode(make_excel_b64("B1")), db_path=core.DB_PATH)

    resp = client.get("/cdc/hang-doi/xuat", params={"commune": "Xã A"})
    from openpyxl import load_workbook
    import tempfile
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "out.xlsx"
        out.write_bytes(resp.content)
        wb = load_workbook(out)
        ws = wb.active
        file_names = [row[2].value for row in ws.iter_rows(min_row=2)]
        assert file_names == ["a.xlsx"]


def test_export_queue_requires_login(client: TestClient):
    resp = client.get("/cdc/hang-doi/xuat", follow_redirects=False)
    assert resp.status_code == 303


def test_queue_page_highlights_duplicate_week_submissions(client: TestClient):
    _login_as(client, core.CDC_ROLE_ADMIN)
    # Nội dung file phải KHÁC nhau (queue_submit chống nộp trùng theo xã+tuần+nội dung giống hệt).
    core.queue_submit("Xã A", "2026-W20", "a1.xlsx", base64.b64decode(make_excel_b64("A1")), db_path=core.DB_PATH)
    core.queue_submit("Xã A", "2026-W20", "a2.xlsx", base64.b64decode(make_excel_b64("A2")), db_path=core.DB_PATH)
    core.queue_submit("Xã B", "2026-W20", "b1.xlsx", base64.b64decode(make_excel_b64("B1")), db_path=core.DB_PATH)

    page = client.get("/cdc/hang-doi")
    assert "cdc-row-warning" in page.text
    assert "Nộp 2 lần/tuần" in page.text


def test_queue_page_does_not_highlight_single_submission(client: TestClient):
    _login_as(client, core.CDC_ROLE_ADMIN)
    core.queue_submit("Xã A", "2026-W20", "a.xlsx", base64.b64decode(make_excel_b64("A1")), db_path=core.DB_PATH)

    page = client.get("/cdc/hang-doi")
    assert "cdc-row-warning" not in page.text


def test_export_queue_includes_duplicate_week_warning_column(client: TestClient):
    _login_as(client, core.CDC_ROLE_ADMIN)
    core.queue_submit("Xã A", "2026-W20", "a1.xlsx", base64.b64decode(make_excel_b64("A1")), db_path=core.DB_PATH)
    core.queue_submit("Xã A", "2026-W20", "a2.xlsx", base64.b64decode(make_excel_b64("A2")), db_path=core.DB_PATH)

    resp = client.get("/cdc/hang-doi/xuat")
    from openpyxl import load_workbook
    import tempfile
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "out.xlsx"
        out.write_bytes(resp.content)
        wb = load_workbook(out)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        assert headers[-1] == "Cảnh báo"
        warnings = [row[-1].value for row in ws.iter_rows(min_row=2)]
        assert warnings.count("Nộp từ 2 lần trở lên trong tuần") == 2


def test_queue_page_has_select_all_checkbox_for_importer(client: TestClient):
    _login_as(client, core.CDC_ROLE_ADMIN)
    page = client.get("/cdc/hang-doi")
    assert "data-batch-select-all" in page.text


def test_queue_list_sorts_by_column_ascending(client: TestClient):
    core.queue_submit("Xã B", "2026-W20", "b.xlsx", base64.b64decode(make_excel_b64("B1")), db_path=core.DB_PATH)
    core.queue_submit("Xã A", "2026-W20", "a.xlsx", base64.b64decode(make_excel_b64("A1")), db_path=core.DB_PATH)
    rows = core.list_import_queue(sort="commune", direction="asc", db_path=core.DB_PATH)
    assert [r["commune"] for r in rows] == ["Xã A", "Xã B"]


def test_queue_list_sorts_by_column_descending(client: TestClient):
    core.queue_submit("Xã B", "2026-W20", "b.xlsx", base64.b64decode(make_excel_b64("B1")), db_path=core.DB_PATH)
    core.queue_submit("Xã A", "2026-W20", "a.xlsx", base64.b64decode(make_excel_b64("A1")), db_path=core.DB_PATH)
    rows = core.list_import_queue(sort="commune", direction="desc", db_path=core.DB_PATH)
    assert [r["commune"] for r in rows] == ["Xã B", "Xã A"]


def test_queue_list_ignores_unknown_sort_column(client: TestClient):
    """Chống SQL injection qua tham số sort — tên cột không nằm trong whitelist phải bị bỏ qua,
    dùng lại thứ tự mặc định thay vì ném lỗi."""
    core.queue_submit("Xã A", "2026-W20", "a.xlsx", base64.b64decode(make_excel_b64("A1")), db_path=core.DB_PATH)
    rows = core.list_import_queue(sort="id; DROP TABLE import_queue;--", db_path=core.DB_PATH)
    assert len(rows) == 1


def test_queue_page_sort_link_reflects_current_column(client: TestClient):
    _login_as(client, core.CDC_ROLE_ADMIN)
    core.queue_submit("Xã B", "2026-W20", "b.xlsx", base64.b64decode(make_excel_b64("B1")), db_path=core.DB_PATH)
    core.queue_submit("Xã A", "2026-W20", "a.xlsx", base64.b64decode(make_excel_b64("A1")), db_path=core.DB_PATH)

    page = client.get("/cdc/hang-doi", params={"sort": "commune", "dir": "asc"})
    assert page.text.index("a.xlsx") < page.text.index("b.xlsx")
    assert "cdc-sort-link--active" in page.text
    assert "dir=desc" in page.text  # bấm lại cùng cột phải đảo chiều


# --- Phân trang (không phá sort) -------------------------------------------------------------

def _seed_many(count: int) -> None:
    for i in range(count):
        core.queue_submit(
            f"Xã {i:03d}", "2026-W20", f"f{i:03d}.xlsx",
            base64.b64decode(make_excel_b64(f"C{i:03d}")), db_path=core.DB_PATH,
        )


def test_queue_page_paginates_at_50_rows(client: TestClient):
    _login_as(client, core.CDC_ROLE_ADMIN)
    _seed_many(60)

    page1 = client.get("/cdc/hang-doi", params={"sort": "commune", "dir": "asc"})
    assert "f000.xlsx" in page1.text
    assert "f049.xlsx" in page1.text
    assert "f050.xlsx" not in page1.text
    assert "60 mục" in page1.text

    page2 = client.get("/cdc/hang-doi", params={"sort": "commune", "dir": "asc", "page": 2})
    assert "f050.xlsx" in page2.text
    assert "f059.xlsx" in page2.text
    assert "f000.xlsx" not in page2.text


def test_queue_pagination_links_preserve_sort(client: TestClient):
    _login_as(client, core.CDC_ROLE_ADMIN)
    _seed_many(60)
    resp = client.get("/cdc/hang-doi", params={"sort": "commune", "dir": "desc"})
    assert "sort=commune" in resp.text and "dir=desc" in resp.text
    assert "page=2" in resp.text


def test_queue_sort_link_does_not_carry_current_page(client: TestClient):
    """Bấm đổi sort phải luôn về trang 1 — link sort không được mang tham số page hiện tại."""
    _login_as(client, core.CDC_ROLE_ADMIN)
    _seed_many(60)
    resp = client.get("/cdc/hang-doi", params={"sort": "commune", "dir": "asc", "page": 2})
    start = resp.text.index('href="/cdc/hang-doi?sort=week')
    end = resp.text.index('"', start + 6)
    assert "page=" not in resp.text[start:end]


def test_admin_can_import_queue_item(client: TestClient):
    _login_as(client, core.CDC_ROLE_ADMIN)
    submitted = core.queue_submit("Xã Gia Viên", "2026-W29", "ds.xlsx", base64.b64decode(make_excel_b64("C1")), db_path=core.DB_PATH)
    csrf = _fresh_csrf(client, "/cdc/hang-doi")
    resp = client.post(f"/cdc/hang-doi/{submitted['queue_id']}/import", data={"csrf_token": csrf}, follow_redirects=False)
    assert resp.status_code == 303
    assert "msg=" in resp.headers["location"]
    item = core.list_import_queue(db_path=core.DB_PATH)[0]
    assert item["status"] == "da_nhap"


def test_viewer_cannot_import(client: TestClient):
    _login_as(client, core.CDC_ROLE_VIEWER)
    submitted = core.queue_submit("Xã Gia Viên", "2026-W29", "ds.xlsx", base64.b64decode(make_excel_b64("C1")), db_path=core.DB_PATH)
    csrf = _fresh_csrf(client, "/cdc/hang-doi")
    resp = client.post(f"/cdc/hang-doi/{submitted['queue_id']}/import", data={"csrf_token": csrf})
    assert resp.status_code == 403


def test_data_operator_cannot_delete(client: TestClient):
    _login_as(client, core.CDC_ROLE_DATA_OPERATOR)
    submitted = core.queue_submit("Xã Gia Viên", "2026-W29", "ds.xlsx", base64.b64decode(make_excel_b64("C1")), db_path=core.DB_PATH)
    csrf = _fresh_csrf(client, "/cdc/hang-doi")
    resp = client.post(f"/cdc/hang-doi/{submitted['queue_id']}/delete", data={"csrf_token": csrf})
    assert resp.status_code == 403
    assert len(core.list_import_queue(db_path=core.DB_PATH)) == 1


def test_admin_can_delete(client: TestClient):
    _login_as(client, core.CDC_ROLE_ADMIN)
    submitted = core.queue_submit("Xã Gia Viên", "2026-W29", "ds.xlsx", base64.b64decode(make_excel_b64("C1")), db_path=core.DB_PATH)
    csrf = _fresh_csrf(client, "/cdc/hang-doi")
    resp = client.post(f"/cdc/hang-doi/{submitted['queue_id']}/delete", data={"csrf_token": csrf}, follow_redirects=False)
    assert resp.status_code == 303
    assert core.list_import_queue(db_path=core.DB_PATH) == []


def test_download_queue_file(client: TestClient):
    _login_as(client, core.CDC_ROLE_VIEWER)
    data = base64.b64decode(make_excel_b64("C1"))
    submitted = core.queue_submit("Xã Gia Viên", "2026-W29", "ds.xlsx", data, db_path=core.DB_PATH)
    resp = client.get(f"/cdc/hang-doi/{submitted['queue_id']}/download")
    assert resp.status_code == 200
    assert resp.content == data
